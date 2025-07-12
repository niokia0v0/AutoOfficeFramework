import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import numpy as np

# Pandas 显示选项
pd.set_option('future.no_silent_downcasting', True)

def process_sales_data(input_file_path):
    """
    处理销售数据CSV文件，生成包含销售总结和各商品销售明细的Excel文件。

    参数:
    input_file_path (str): 输入的CSV文件的完整路径。

    返回:
    str: 生成的Excel文件的完整路径。
    """

    # ---- 1. 初始化路径和读取数据 ----
    input_dir = os.path.dirname(input_file_path)
    base_name = os.path.basename(input_file_path)
    file_name_without_ext = os.path.splitext(base_name)[0]
    output_file_path = os.path.join(input_dir, f"{file_name_without_ext}_output_JD.xlsx")

    DETAIL_SHEET_COLUMNS = ['订单编号', '父单号', '订单状态', '订单下单时间', '订单完成时间',
                            '售后服务单号', '售后退款时间', '商品编号', '商品名称',
                            '商品数量', '扣点类型', '佣金比例', '费用名称', '应结金额',
                            '收支方向', '结算状态', '预计结算时间', '账单生成时间',
                            '到账时间', '商户订单号', '资金动账备注', '费用项含义', '备注',
                            '留用时间', '费用说明']

    df = pd.read_csv(input_file_path, dtype=str, na_values=['--'], keep_default_na=True)

    # ---- 2. 数据清洗和预处理 ----
    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = df[col].str.strip()
            df[col] = df[col].replace(['--', '', 'None', 'nan'], np.nan, regex=False)

    df['应结金额'] = pd.to_numeric(df['应结金额'], errors='coerce').fillna(0)
    df['商品数量'] = pd.to_numeric(df['商品数量'], errors='coerce').fillna(0)
    if '佣金比例' in df.columns:
        df['佣金比例'] = pd.to_numeric(df['佣金比例'], errors='coerce')

    all_completed_df = df[df['订单状态'] == '已完成'].copy()
    for col_name_to_check in ['售后服务单号', '售后退款时间']:
        if col_name_to_check not in all_completed_df.columns:
            all_completed_df[col_name_to_check] = np.nan

    completed_with_product_id_df = all_completed_df[all_completed_df['商品编号'].notna()].copy()

    # ---- 3. 按商品编号汇总销售和支出信息 ----
    product_summary_details = {}
    for product_id_raw, group in completed_with_product_id_df.groupby('商品编号'):
        product_id = str(product_id_raw)
        product_name_series = group['商品名称'].dropna()
        product_name = product_name_series.iloc[0] if not product_name_series.empty else "未知商品"

        sales_group_full_cols = group[(group['费用名称'] == '货款') & (group['收支方向'] == '收入')].copy()
        sales_group_for_detail_sheet = sales_group_full_cols[[col for col in DETAIL_SHEET_COLUMNS if col in sales_group_full_cols.columns]].copy()
        sales_quantity = sales_group_full_cols['商品数量'].sum()
        sales_amount = sales_group_full_cols['应结金额'].sum()

        returns_group_full_cols = group[
            (group['费用名称'] == '货款') &
            (group['收支方向'] == '支出') &
            (group['售后服务单号'].notna()) &
            (group['售后服务单号'] != '')
        ].copy()
        returns_group_for_detail_sheet = returns_group_full_cols[[col for col in DETAIL_SHEET_COLUMNS if col in returns_group_full_cols.columns]].copy()
        return_quantity = returns_group_full_cols['商品数量'].sum()
        return_amount_negative = returns_group_full_cols['应结金额'].sum()

        product_expenses_group = group[group['收支方向'] == '支出']
        total_commission = product_expenses_group[product_expenses_group['费用名称'] == '佣金']['应结金额'].sum()
        total_transaction_fee = product_expenses_group[product_expenses_group['费用名称'] == '交易服务费']['应结金额'].sum()
        total_ad_commission = product_expenses_group[product_expenses_group['费用名称'] == '广告联合活动降扣佣金']['应结金额'].sum()
        total_jingdou = product_expenses_group[product_expenses_group['费用名称'] == '京豆']['应结金额'].sum()

        relevant_order_ids = group['订单编号'].unique()
        orders_containing_product = all_completed_df[all_completed_df['订单编号'].isin(relevant_order_ids)]
        total_product_insurance_for_orders = orders_containing_product[
            (orders_containing_product['费用名称'] == '商品保险服务费') &
            (orders_containing_product['收支方向'] == '支出')
        ]['应结金额'].sum()
        total_freight_insurance_for_orders = orders_containing_product[
            (orders_containing_product['费用名称'] == '运费保险服务费') &
            (orders_containing_product['收支方向'] == '支出')
        ]['应结金额'].sum()
        total_product_related_expenses = (total_commission + total_transaction_fee +
                                          total_ad_commission + total_jingdou +
                                          total_product_insurance_for_orders +
                                          total_freight_insurance_for_orders)

        product_summary_details[product_id] = {
            'name': product_name,
            'sales_quantity': sales_quantity,
            'sales_amount': sales_amount,
            'return_quantity': return_quantity,
            'return_amount_negative': return_amount_negative,
            'commission': total_commission,
            'transaction_fee': total_transaction_fee,
            'ad_commission': total_ad_commission,
            'jingdou': total_jingdou,
            'product_insurance': total_product_insurance_for_orders,
            'freight_insurance': total_freight_insurance_for_orders,
            'total_product_expenses': total_product_related_expenses,
            'sales_detail_data': sales_group_for_detail_sheet,
            'returns_detail_data': returns_group_for_detail_sheet
        }

    # ---- 4. 创建Excel工作簿并写入销售总结页 ----
    wb = Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "销售总结"
    summary_header = [
        "商品编号", "商品名称", "销售数量", "销售额",
        "佣金支出", "交易服务费支出", "广告降扣支出", "京豆支出",
        "商品保险费支出", "运费保险费支出", "产品总支出"
    ]
    summary_sheet.append(summary_header)
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    total_sales_qty_no_refund = 0
    total_sales_amt_no_refund = 0
    total_commission_no_refund = 0
    total_txn_fee_no_refund = 0
    total_ad_comm_no_refund = 0
    total_jingdou_no_refund = 0
    total_prod_ins_no_refund = 0
    total_freight_ins_no_refund = 0
    total_prod_exp_no_refund = 0
    grand_total_return_quantity_for_summary = 0
    grand_total_return_amount_for_summary = 0

    sorted_summary_keys = sorted(product_summary_details.keys())
    for prod_id in sorted_summary_keys:
        item = product_summary_details[prod_id]
        summary_sheet.append([
            prod_id, item['name'],
            item['sales_quantity'], item['sales_amount'],
            item['commission'], item['transaction_fee'], item['ad_commission'], item['jingdou'],
            item['product_insurance'], item['freight_insurance'], item['total_product_expenses']
        ])
        total_sales_qty_no_refund += item['sales_quantity']
        total_sales_amt_no_refund += item['sales_amount']
        total_commission_no_refund += item['commission']
        total_txn_fee_no_refund += item['transaction_fee']
        total_ad_comm_no_refund += item['ad_commission']
        total_jingdou_no_refund += item['jingdou']
        total_prod_ins_no_refund += item['product_insurance']
        total_freight_ins_no_refund += item['freight_insurance']
        total_prod_exp_no_refund += item['total_product_expenses']
        grand_total_return_quantity_for_summary += item['return_quantity']
        grand_total_return_amount_for_summary += item['return_amount_negative']

    current_row = summary_sheet.max_row + 1
    
    # 1. 总计 (不含退款)
    summary_sheet.cell(row=current_row, column=1, value="总计 (不含退款)").font = Font(bold=True)
    summary_sheet.cell(row=current_row, column=3, value=total_sales_qty_no_refund).font = Font(bold=True)
    summary_sheet.cell(row=current_row, column=4, value=total_sales_amt_no_refund).font = Font(bold=True)
    summary_sheet.cell(row=current_row, column=5, value=total_commission_no_refund).font = Font(bold=True)
    summary_sheet.cell(row=current_row, column=6, value=total_txn_fee_no_refund).font = Font(bold=True)
    summary_sheet.cell(row=current_row, column=7, value=total_ad_comm_no_refund).font = Font(bold=True)
    summary_sheet.cell(row=current_row, column=8, value=total_jingdou_no_refund).font = Font(bold=True)
    summary_sheet.cell(row=current_row, column=9, value=total_prod_ins_no_refund).font = Font(bold=True)
    summary_sheet.cell(row=current_row, column=10, value=total_freight_ins_no_refund).font = Font(bold=True)
    summary_sheet.cell(row=current_row, column=11, value=total_prod_exp_no_refund).font = Font(bold=True)
    for c_idx in range(3, len(summary_header) + 1): # 从第3列开始格式化数值
        cell = summary_sheet.cell(row=current_row, column=c_idx)
        if c_idx == 3: cell.number_format = '#,##0'
        else: cell.number_format = '#,##0.00'
    
    current_row += 1
    summary_sheet.append([]) 
    current_row += 1

    # 2. 退款商品明细及 "总计退款" 行
    refund_items_exist = False
    temp_return_qty_sum = 0
    temp_return_amt_sum = 0
    
    # 先写入退款的商品条目
    for prod_id in sorted_summary_keys:
        item = product_summary_details[prod_id]
        if item['return_quantity'] > 0:
            refund_items_exist = True
            summary_sheet.cell(row=current_row, column=1, value=prod_id) 
            summary_sheet.cell(row=current_row, column=2, value=item['name']) 
            summary_sheet.cell(row=current_row, column=3, value=item['return_quantity']).number_format = '#,##0'
            summary_sheet.cell(row=current_row, column=4, value=item['return_amount_negative']).number_format = '#,##0.00'
            current_row +=1
            temp_return_qty_sum += item['return_quantity'] # 临时累加，用于总计退款行
            temp_return_amt_sum += item['return_amount_negative']
            
    # 写入 "总计退款" 行
    summary_sheet.cell(row=current_row, column=1, value="总计退款").font = Font(bold=True)
    summary_sheet.cell(row=current_row, column=3, value=temp_return_qty_sum).font = Font(bold=True)
    summary_sheet.cell(row=current_row, column=3).number_format = '#,##0'
    summary_sheet.cell(row=current_row, column=4, value=temp_return_amt_sum).font = Font(bold=True)
    summary_sheet.cell(row=current_row, column=4).number_format = '#,##0.00'

    # 3. 总计 (计算退款)
    current_row += 1
    summary_sheet.append([]) 
    current_row += 1
    final_total_sales_qty = total_sales_qty_no_refund - grand_total_return_quantity_for_summary
    final_total_sales_amt = total_sales_amt_no_refund + grand_total_return_amount_for_summary

    summary_sheet.cell(row=current_row, column=1, value="总计 (计算退款)").font = Font(bold=True)
    summary_sheet.cell(row=current_row, column=3, value=final_total_sales_qty).font = Font(bold=True)
    summary_sheet.cell(row=current_row, column=3).number_format = '#,##0'
    summary_sheet.cell(row=current_row, column=4, value=final_total_sales_amt).font = Font(bold=True)
    summary_sheet.cell(row=current_row, column=4).number_format = '#,##0.00'
    # 其他支出列在此行中留空


    # 为数据行（商品行）设置格式
    first_total_row = 0
    for r_idx in range(2, summary_sheet.max_row + 1):
        if summary_sheet.cell(row=r_idx, column=1).value == "总计 (不含退款)":
            first_total_row = r_idx
            break
    if first_total_row > 0:
        for row_idx_num in range(2, first_total_row): 
            summary_sheet.cell(row=row_idx_num, column=3).number_format = '#,##0' 
            for col_idx_format in range(4, len(summary_header) + 1): 
                summary_sheet.cell(row=row_idx_num, column=col_idx_format).number_format = '#,##0.00'
    
    summary_sheet.column_dimensions['A'].width = 20
    summary_sheet.column_dimensions['B'].width = 60
    summary_sheet.column_dimensions['C'].width = 12
    summary_sheet.column_dimensions['D'].width = 15
    for col_letter_idx in range(ord('E'), ord('L') + 1): # E to K for new header
         if col_letter_idx <= ord('K'): # K is the last column "产品总支出"
            summary_sheet.column_dimensions[chr(col_letter_idx)].width = 18

    # ---- 5. 为每个商品创建并写入详情页 ---- (此部分代码不变)
    # ... (与上一版相同的详情页创建逻辑) ...
    for product_id_str in sorted_summary_keys:
        product_info = product_summary_details[product_id_str]
        
        sales_df_for_detail_orig = product_info['sales_detail_data'].copy() 
        returns_df_for_detail_orig = product_info['returns_detail_data'].copy() 

        if sales_df_for_detail_orig.empty and returns_df_for_detail_orig.empty:
            continue
        
        clean_product_name = re.sub(r'[\\/\*\[\]\:?]', '_', str(product_info['name']))
        potential_sheet_name = f"{product_id_str}_{clean_product_name}" 
        if len(potential_sheet_name) > 31:
            id_len = len(product_id_str)
            name_part_len = 31 - id_len - 1 
            sheet_name = f"{product_id_str}_{clean_product_name[:name_part_len]}" if name_part_len > 5 else potential_sheet_name[:31]
        else:
            sheet_name = potential_sheet_name
        product_sheet = wb.create_sheet(sheet_name) 

        header_written_for_sheet = False 

        def write_section_to_sheet(df_section_orig, is_header_needed):
            nonlocal header_written_for_sheet 
            if df_section_orig.empty: 
                return

            df_to_write = df_section_orig.copy() 

            for col_to_ensure in DETAIL_SHEET_COLUMNS:
                if col_to_ensure not in df_to_write.columns:
                    df_to_write[col_to_ensure] = np.nan 
            
            df_to_write = df_to_write[DETAIL_SHEET_COLUMNS] 

            numeric_cols_in_detail = ['商品数量', '应结金额', '佣金比例'] 
            for col_name in df_to_write.columns:
                if col_name in numeric_cols_in_detail:
                    if col_name == '佣金比例': 
                         df_to_write[col_name] = pd.to_numeric(df_to_write[col_name], errors='coerce')
                    else: 
                        df_to_write[col_name] = pd.to_numeric(df_to_write[col_name], errors='coerce').fillna(0)
                else: 
                    df_to_write[col_name] = df_to_write[col_name].fillna('').astype(str)

            if is_header_needed and not header_written_for_sheet:
                product_sheet.append(DETAIL_SHEET_COLUMNS) 
                for cell in product_sheet[product_sheet.max_row]: 
                    cell.font = Font(bold=True)
                header_written_for_sheet = True 
            
            for row_data_tuple in df_to_write.itertuples(index=False):
                product_sheet.append(list(row_data_tuple))

        write_section_to_sheet(sales_df_for_detail_orig, True) 

        if not sales_df_for_detail_orig.empty: 
            detail_sales_total_row_idx = product_sheet.max_row + 1
            product_sheet.cell(row=detail_sales_total_row_idx, column=1, value="销售总计").font = Font(bold=True)
            
            if '商品数量' in DETAIL_SHEET_COLUMNS:
                qty_col_idx = DETAIL_SHEET_COLUMNS.index('商品数量') + 1
                cell_qty = product_sheet.cell(row=detail_sales_total_row_idx, column=qty_col_idx, value=product_info['sales_quantity'])
                cell_qty.font = Font(bold=True); cell_qty.number_format = '#,##0'
            
            if '应结金额' in DETAIL_SHEET_COLUMNS:
                amt_col_idx = DETAIL_SHEET_COLUMNS.index('应结金额') + 1
                cell_amt = product_sheet.cell(row=detail_sales_total_row_idx, column=amt_col_idx, value=product_info['sales_amount'])
                cell_amt.font = Font(bold=True); cell_amt.number_format = '#,##0.00'

        if not returns_df_for_detail_orig.empty: 
            product_sheet.append([]) 
            write_section_to_sheet(returns_df_for_detail_orig, not header_written_for_sheet)

        if header_written_for_sheet: 
            for current_col_idx, current_column_title in enumerate(DETAIL_SHEET_COLUMNS, 1):
                column_letter_val_detail = get_column_letter(current_col_idx)
                max_len = len(str(current_column_title))
                for row_num_val_detail in range(1, product_sheet.max_row + 1):
                    cell_val_obj_detail = product_sheet.cell(row=row_num_val_detail, column=current_col_idx).value
                    if cell_val_obj_detail is not None:
                        if isinstance(cell_val_obj_detail, (int, float)) and cell_val_obj_detail != 0:
                            if current_column_title == '商品数量':
                                max_len = max(max_len, len(f"{cell_val_obj_detail:,}"))
                            elif current_column_title == '应结金额':
                                 max_len = max(max_len, len(f"{cell_val_obj_detail:,.2f}"))
                            elif current_column_title == '佣金比例' and not pd.isna(cell_val_obj_detail): 
                                max_len = max(max_len, len(f"{cell_val_obj_detail:.4f}"))
                            else: 
                                max_len = max(max_len, len(str(cell_val_obj_detail)))
                        else: 
                             max_len = max(max_len, len(str(cell_val_obj_detail)))
                
                adjusted_width = min(max(max_len + 3, 10), 60)
                if current_column_title == "商品名称":
                    adjusted_width = min(max(max_len + 3, 40), 70)
                elif current_column_title in ['备注', '费用说明', '费用项含义']:
                    adjusted_width = min(max(max_len + 3, 20), 70)
                product_sheet.column_dimensions[column_letter_val_detail].width = adjusted_width

    # ---- 6. 确保 "销售总结" 工作表为第一个 ----
    if wb.sheetnames[0] != "销售总结" and "销售总结" in wb.sheetnames:
        summary_s = wb["销售总结"]
        wb._sheets.remove(summary_s)
        wb._sheets.insert(0, summary_s)

    # ---- 7. 保存Excel工作簿 ----
    wb.save(output_file_path)
    return output_file_path

# ---- 主程序入口 ----
if __name__ == "__main__":
    directory_path = r"F:\étude\Ecole\E4\E4stage\E4stageProjet\samples\JD"#待处理文件所在目录
    csv_filename = "订单结算明细对账_2025-05-01_2025-05-31 (1).csv"#待处理文件名
    input_file = os.path.join(directory_path, csv_filename)
    
    if not os.path.exists(input_file):
        print(f"错误: 输入文件未找到于 '{os.path.abspath(input_file)}'")
    elif not input_file.lower().endswith(".csv"):
        print(f"错误: 输入文件 '{os.path.basename(input_file)}' 不是一个CSV文件。")
    else:
        try:
            print(f"开始处理文件: {input_file}")
            output_file = process_sales_data(input_file)
            print(f"处理完成。输出文件位于: {output_file}")
        except FileNotFoundError:
             print(f"错误: 输入文件未找到于 '{os.path.abspath(input_file)}'")
        except pd.errors.EmptyDataError:
            print(f"错误: 输入文件 '{os.path.basename(input_file)}' 为空或格式不正确。")
        except Exception as e:
            print(f"处理过程中发生错误: {e}")
            import traceback
            traceback.print_exc()