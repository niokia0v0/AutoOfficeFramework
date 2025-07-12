import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import numpy as np

# Pandas 显示选项
pd.set_option('future.no_silent_downcasting', True)

# --- 配置项 ---
# 天猫原始列名映射
TMALL_COL_SUB_ORDER_ID = '子订单编号'         # A列
TMALL_COL_MAIN_ORDER_ID = '主订单编号'        # B列
TMALL_COL_PRODUCT_NAME = '商品标题'         # C列
TMALL_COL_UNIT_PRICE = '商品价格'           # D列
TMALL_COL_QUANTITY = '购买数量'             # E列
TMALL_COL_PRODUCT_ATTRIBUTES = '商品属性'     # G列 (现在要在详情页使用)
TMALL_COL_ORDER_STATUS = '订单状态'         # J列 (核心)
TMALL_COL_SELLER_SKU = '商家编码'           # K列 (不再直接用于详情页主要展示，但常量保留以防他用)
TMALL_COL_ACTUAL_PAYMENT = '买家实付金额'     # M列 (核心)
TMALL_COL_REFUND_STATUS = '退款状态'        # N列
TMALL_COL_REFUND_AMOUNT = '退款金额'          # O列 (核心)
TMALL_COL_ORDER_CREATE_TIME = '订单创建时间'  # P列
TMALL_COL_ORDER_PAY_TIME = '订单付款时间'       # Q列
TMALL_COL_PRODUCT_ID = '商品ID'             # R列 (核心)
TMALL_COL_SHIPPING_TIME = '发货时间'          # U列
TMALL_COL_LOGISTICS_NO = '物流单号'           # V列
TMALL_COL_LOGISTICS_COMPANY = '物流公司'      # W列
# 可选列
TMALL_COL_EXTERNAL_SYSTEM_ID = '外部系统编号' # F列
TMALL_COL_PACKAGE_INFO = '套餐信息'           # H列
TMALL_COL_CONTACT_REMARK = '联系方式备注'     # I列
TMALL_COL_BUYER_PAYABLE_GOODS = '买家应付货款'# L列
TMALL_COL_SELLER_REMARK = '商家备注'          # S列
TMALL_COL_BUYER_MESSAGE = '主订单买家留言'    # T列

# 订单状态常量
STATUS_TRADE_SUCCESS = '交易成功' # J列 '订单状态' 中表示交易成功的确切文本

# 输出到详情页的列定义 (已将“商家编码”替换为“商品属性”，并保持时间列调整后的顺序)
DETAIL_SHEET_COLUMNS_TMALL = [
    '订单编号',
    '子订单编号',
    '订单状态',
    '退款状态',
    '商品编号',
    '商品名称',
    '商品属性',  # <--- 修改处
    '商品价格',
    '商品数量',
    '应结金额',
    '订单创建时间',
    '订单付款时间',
    '发货时间',
    '物流单号',
    '物流公司',
]

def process_tmall_sales_data(input_file_path):
    """
    处理天猫订单导出的XLSX文件，生成包含销售总结和各商品销售明细的Excel文件。
    总结页和详情页均按“收入”和“支出”逻辑进行分区。
    总结页的“净销售数量” = (所有订单购买数) - (所有未成功订单购买数)。
    详情页的“商家编码”已替换为“商品属性”，时间列顺序已调整。
    “买家实付款总额”行现在显示“交易成功”订单的买家实付金额总和。
    """
    # --- 1. 初始化路径和读取数据 ---
    input_dir = os.path.dirname(input_file_path)
    base_name = os.path.basename(input_file_path)
    file_name_without_ext = os.path.splitext(base_name)[0]
    output_file_path = os.path.join(input_dir, f"{file_name_without_ext}_output_TM.xlsx")

    try:
        try:
            df_original = pd.read_excel(input_file_path, dtype=str)
        except UnicodeDecodeError:
            df_original = pd.read_excel(input_file_path, dtype=str, engine='openpyxl')
    except FileNotFoundError:
        print(f"错误: 输入文件未找到于 '{os.path.abspath(input_file_path)}'")
        return None
    except Exception as e:
        print(f"读取Excel文件 '{input_file_path}' 失败: {e}")
        return None

    if df_original.empty:
        print(f"错误: 输入文件 '{base_name}' 为空或未能正确读取数据。")
        return None

    # --- 2. 数据清洗和预处理 ---
    df_original.columns = df_original.columns.str.strip()
    for col in df_original.columns:
        if df_original[col].dtype == 'object':
            df_original[col] = df_original[col].str.strip()
            df_original[col] = df_original[col].replace(
                ['--', '', 'None', 'nan', '#NULL!', None], np.nan, regex=False
            )

    critical_logic_cols_check = {
        TMALL_COL_PRODUCT_ID: "商品ID (R列)", TMALL_COL_ORDER_STATUS: "订单状态 (J列)",
        TMALL_COL_QUANTITY: "购买数量 (E列)", TMALL_COL_ACTUAL_PAYMENT: "买家实付金额 (M列)",
        TMALL_COL_REFUND_AMOUNT: "退款金额 (O列)",
        TMALL_COL_PRODUCT_ATTRIBUTES: "商品属性 (G列)"
    }
    for col_const, col_display_name in critical_logic_cols_check.items():
        if col_const not in df_original.columns:
            print(f"错误: 核心逻辑或详情页所需列 '{col_display_name}' (脚本常量名: {col_const}) "
                  f"在输入文件 '{base_name}' 的表头中未找到。请检查脚本顶部的列名配置。脚本无法继续。")
            return None

    numeric_cols_to_convert_map = {
        TMALL_COL_QUANTITY: 0, TMALL_COL_ACTUAL_PAYMENT: 0.0,
        TMALL_COL_REFUND_AMOUNT: 0.0, TMALL_COL_UNIT_PRICE: 0.0
    }
    for col, default_fill_value in numeric_cols_to_convert_map.items():
        if col in df_original.columns:
            df_original[col] = pd.to_numeric(df_original[col], errors='coerce').fillna(default_fill_value)
        elif col == TMALL_COL_UNIT_PRICE:
             print(f"警告: 列 '{col}' ('商品价格') 在输入文件中未找到，将创建此列并使用默认值 {default_fill_value} 填充。")
             df_original[col] = default_fill_value

    df_original[TMALL_COL_PRODUCT_ID] = df_original[TMALL_COL_PRODUCT_ID].astype(str).replace('nan', np.nan)
    df_processed = df_original[df_original[TMALL_COL_PRODUCT_ID].notna()].copy()
    if df_processed.empty:
        print(f"数据中没有找到包含有效商品ID ('{TMALL_COL_PRODUCT_ID}') 的行。无法生成报告。")
        return None

    # --- 新增计算: "交易成功"订单的买家实付总额 ---
    # 确保 TMALL_COL_ORDER_STATUS 和 TMALL_COL_ACTUAL_PAYMENT 存在于 df_processed
    actual_payment_for_successful_trades = 0.0
    if TMALL_COL_ORDER_STATUS in df_processed.columns and TMALL_COL_ACTUAL_PAYMENT in df_processed.columns:
        successful_trades_df = df_processed[df_processed[TMALL_COL_ORDER_STATUS] == STATUS_TRADE_SUCCESS]
        actual_payment_for_successful_trades = successful_trades_df[TMALL_COL_ACTUAL_PAYMENT].sum()
    else:
        print(f"警告: 无法计算“交易成功”订单的实付款总额，因为列'{TMALL_COL_ORDER_STATUS}'或'{TMALL_COL_ACTUAL_PAYMENT}'不存在。")


    # --- 3. 按商品ID汇总信息，用于总结页和详情页 ---
    product_data_map = {}

    for product_id_value, group_df in df_processed.groupby(TMALL_COL_PRODUCT_ID):
        product_id_str_key = str(product_id_value)
        product_name_series_data = group_df[TMALL_COL_PRODUCT_NAME].dropna() if TMALL_COL_PRODUCT_NAME in group_df else pd.Series([])
        product_name_str = product_name_series_data.iloc[0] if not product_name_series_data.empty else "未知商品"

        all_orders_in_group = group_df.copy()
        income_total_quantity_per_product = all_orders_in_group[TMALL_COL_QUANTITY].sum()
        # "总计收入"中的每项商品收入依然是基于该商品所有订单的实付款
        income_total_amount_per_product = all_orders_in_group[TMALL_COL_ACTUAL_PAYMENT].sum()


        non_successful_orders_in_group = group_df[group_df[TMALL_COL_ORDER_STATUS] != STATUS_TRADE_SUCCESS].copy()
        expenditure_total_quantity_per_product = non_successful_orders_in_group[TMALL_COL_QUANTITY].sum()
        expenditure_total_amount_per_product = -non_successful_orders_in_group[TMALL_COL_REFUND_AMOUNT].sum()

        def format_df_for_detail_sheet(source_df_input, prod_id_for_detail, prod_name_for_detail,
                                     amount_source_col_name, make_amount_negative=False):
            if source_df_input.empty:
                return pd.DataFrame(columns=DETAIL_SHEET_COLUMNS_TMALL)

            detail_df_formatted = pd.DataFrame()
            detail_df_formatted['订单编号'] = source_df_input.get(TMALL_COL_MAIN_ORDER_ID)
            detail_df_formatted['子订单编号'] = source_df_input.get(TMALL_COL_SUB_ORDER_ID)
            detail_df_formatted['订单状态'] = source_df_input.get(TMALL_COL_ORDER_STATUS)
            detail_df_formatted['退款状态'] = source_df_input.get(TMALL_COL_REFUND_STATUS)
            detail_df_formatted['订单创建时间'] = source_df_input.get(TMALL_COL_ORDER_CREATE_TIME)
            detail_df_formatted['订单付款时间'] = source_df_input.get(TMALL_COL_ORDER_PAY_TIME)
            detail_df_formatted['发货时间'] = source_df_input.get(TMALL_COL_SHIPPING_TIME)
            detail_df_formatted['商品编号'] = prod_id_for_detail
            detail_df_formatted['商品名称'] = prod_name_for_detail
            detail_df_formatted['商品属性'] = source_df_input.get(TMALL_COL_PRODUCT_ATTRIBUTES)
            detail_df_formatted['商品价格'] = source_df_input.get(TMALL_COL_UNIT_PRICE)
            detail_df_formatted['商品数量'] = source_df_input.get(TMALL_COL_QUANTITY)

            amount_values = source_df_input.get(amount_source_col_name, 0.0)
            detail_df_formatted['应结金额'] = -amount_values if make_amount_negative else amount_values

            detail_df_formatted['物流单号'] = source_df_input.get(TMALL_COL_LOGISTICS_NO)
            detail_df_formatted['物流公司'] = source_df_input.get(TMALL_COL_LOGISTICS_COMPANY)

            final_detail_df = pd.DataFrame(columns=DETAIL_SHEET_COLUMNS_TMALL)
            for col_name in DETAIL_SHEET_COLUMNS_TMALL:
                if col_name in detail_df_formatted.columns:
                    final_detail_df[col_name] = detail_df_formatted[col_name]
                else:
                    if col_name in ['商品价格', '商品数量', '应结金额']:
                        final_detail_df[col_name] = 0 if col_name == '商品数量' else 0.0
                    else:
                        final_detail_df[col_name] = ''
            return final_detail_df


        detail_income_section_df = format_df_for_detail_sheet(
            all_orders_in_group, product_id_str_key, product_name_str,
            TMALL_COL_ACTUAL_PAYMENT, make_amount_negative=False
        )
        detail_expenditure_section_df = format_df_for_detail_sheet(
            non_successful_orders_in_group, product_id_str_key, product_name_str,
            TMALL_COL_REFUND_AMOUNT, make_amount_negative=True
        )

        if not detail_income_section_df.empty or not detail_expenditure_section_df.empty:
            product_data_map[product_id_str_key] = {
                'name': product_name_str,
                'income_total_quantity': income_total_quantity_per_product,
                'income_total_amount': income_total_amount_per_product,
                'expenditure_total_quantity': expenditure_total_quantity_per_product,
                'expenditure_total_amount': expenditure_total_amount_per_product,
                'detail_income_df': detail_income_section_df,
                'detail_expenditure_df': detail_expenditure_section_df,
            }

    # --- 4. 创建Excel工作簿并写入销售总结页 ---
    wb = Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "销售总结"

    summary_income_header = ["商品编号", "商品名称", "总商品数量", "总销售额(收入)"]
    summary_expenditure_header = ["商品编号", "商品名称", "未成功订单商品数量", "总退款额(支出)"]
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')

    current_row_summary = 1

    summary_sheet.cell(row=current_row_summary, column=1, value="各商品收入汇总").font = bold_font
    current_row_summary += 1
    summary_sheet.append(summary_income_header)
    for cell in summary_sheet[current_row_summary]: cell.font = bold_font; cell.alignment = center_alignment
    current_row_summary += 1

    grand_total_income_qty = 0
    # grand_total_income_amt 代表所有（有商品ID）订单的实付金额总和，用于“总计收入”行
    grand_total_income_amt = 0
    sorted_product_ids = sorted(product_data_map.keys())

    for prod_id in sorted_product_ids:
        item_data = product_data_map[prod_id]
        summary_sheet.append([
            prod_id, item_data['name'],
            item_data['income_total_quantity'], item_data['income_total_amount']
        ])
        grand_total_income_qty += item_data['income_total_quantity']
        grand_total_income_amt += item_data['income_total_amount'] # This sums up all actual payments for products
        summary_sheet.cell(row=current_row_summary, column=3).number_format = '#,##0'
        summary_sheet.cell(row=current_row_summary, column=4).number_format = '#,##0.00'
        current_row_summary += 1

    # "总计收入" 行显示的是所有商品（即所有有product_id的订单）的买家实付总额
    summary_sheet.cell(row=current_row_summary, column=1, value="总计收入").font = bold_font
    summary_sheet.cell(row=current_row_summary, column=3, value=grand_total_income_qty).font = bold_font
    summary_sheet.cell(row=current_row_summary, column=3).number_format = '#,##0'
    summary_sheet.cell(row=current_row_summary, column=4, value=grand_total_income_amt).font = bold_font
    summary_sheet.cell(row=current_row_summary, column=4).number_format = '#,##0.00'
    current_row_summary += 1
    summary_sheet.append([])
    current_row_summary += 1

    summary_sheet.cell(row=current_row_summary, column=1, value="各商品支出汇总").font = bold_font
    current_row_summary += 1
    summary_sheet.append(summary_expenditure_header)
    for cell in summary_sheet[current_row_summary]: cell.font = bold_font; cell.alignment = center_alignment
    current_row_summary += 1

    grand_total_expenditure_qty = 0
    grand_total_expenditure_amt = 0

    for prod_id in sorted_product_ids:
        item_data = product_data_map[prod_id]
        if item_data['expenditure_total_quantity'] > 0 or item_data['expenditure_total_amount'] != 0:
            summary_sheet.append([
                prod_id, item_data['name'],
                item_data['expenditure_total_quantity'], item_data['expenditure_total_amount']
            ])
            grand_total_expenditure_qty += item_data['expenditure_total_quantity']
            grand_total_expenditure_amt += item_data['expenditure_total_amount']
            summary_sheet.cell(row=current_row_summary, column=3).number_format = '#,##0'
            summary_sheet.cell(row=current_row_summary, column=4).number_format = '#,##0.00'
            current_row_summary += 1

    summary_sheet.cell(row=current_row_summary, column=1, value="总计支出").font = bold_font
    summary_sheet.cell(row=current_row_summary, column=3, value=grand_total_expenditure_qty).font = bold_font
    summary_sheet.cell(row=current_row_summary, column=3).number_format = '#,##0'
    summary_sheet.cell(row=current_row_summary, column=4, value=grand_total_expenditure_amt).font = bold_font
    summary_sheet.cell(row=current_row_summary, column=4).number_format = '#,##0.00'
    current_row_summary += 1
    summary_sheet.append([])
    current_row_summary += 1

    net_total_quantity = grand_total_income_qty - grand_total_expenditure_qty
    net_total_amount = grand_total_income_amt + grand_total_expenditure_amt

    summary_sheet.cell(row=current_row_summary, column=1, value="净总计").font = bold_font
    summary_sheet.cell(row=current_row_summary, column=3, value=net_total_quantity).font = bold_font
    summary_sheet.cell(row=current_row_summary, column=3).number_format = '#,##0'
    summary_sheet.cell(row=current_row_summary, column=4, value=net_total_amount).font = bold_font
    summary_sheet.cell(row=current_row_summary, column=4).number_format = '#,##0.00'
    current_row_summary += 1

    # --- 修改: 使用 actual_payment_for_successful_trades ---
    summary_sheet.cell(row=current_row_summary, column=1, value="买家实付款总额(交易成功订单)").font = bold_font
    summary_sheet.cell(row=current_row_summary, column=4, value=actual_payment_for_successful_trades).font = bold_font
    summary_sheet.cell(row=current_row_summary, column=4).number_format = '#,##0.00'
    # --- 修改结束 ---

    summary_sheet.column_dimensions['A'].width = 35 # 调整A列宽度以适应新标题
    summary_sheet.column_dimensions['B'].width = 60
    summary_sheet.column_dimensions['C'].width = 18
    summary_sheet.column_dimensions['D'].width = 20

    # --- 5. 为每个商品创建并写入详情页 ---
    # (这部分代码保持不变)
    for product_id_str_key in sorted_product_ids:
        product_info_item = product_data_map[product_id_str_key]
        detail_income_df_data = product_info_item['detail_income_df']
        detail_expenditure_df_data = product_info_item['detail_expenditure_df']

        if detail_income_df_data.empty and detail_expenditure_df_data.empty:
            continue

        clean_product_name_str = re.sub(r'[\\/\*\[\]\:?]', '_', str(product_info_item['name']))
        potential_sheet_name_str = f"{product_id_str_key}_{clean_product_name_str}"
        sheet_name_final = potential_sheet_name_str[:31] if len(potential_sheet_name_str) > 31 else potential_sheet_name_str
        try:
            product_detail_sheet = wb.create_sheet(sheet_name_final)
        except:
            product_detail_sheet = wb.create_sheet(potential_sheet_name_str[:28] + "...")

        header_written_for_this_sheet = False

        def write_section_to_sheet_detail(current_sheet_obj, df_section_data,
                                         is_header_needed_flag, section_title_str=None):
            nonlocal header_written_for_this_sheet

            if df_section_data.empty and not section_title_str :
                return

            df_to_write_to_excel = df_section_data.copy()
            df_to_write_to_excel = df_to_write_to_excel.reindex(columns=DETAIL_SHEET_COLUMNS_TMALL, fill_value='')


            for col_name_fmt in df_to_write_to_excel.columns:
                if col_name_fmt not in ['商品价格', '商品数量', '应结金额']:
                    df_to_write_to_excel[col_name_fmt] = df_to_write_to_excel[col_name_fmt].fillna('').astype(str)

            if is_header_needed_flag and not header_written_for_this_sheet:
                current_sheet_obj.append(DETAIL_SHEET_COLUMNS_TMALL)
                for cell_header_obj in current_sheet_obj[current_sheet_obj.max_row]:
                    cell_header_obj.font = bold_font
                    cell_header_obj.alignment = center_alignment
                header_written_for_this_sheet = True

            for _, row_data_item in df_to_write_to_excel.iterrows():
                 current_sheet_obj.append(row_data_item.tolist())

            if section_title_str:
                total_row_idx_for_section = current_sheet_obj.max_row + 1
                current_sheet_obj.cell(row=total_row_idx_for_section, column=1, value=section_title_str).font = bold_font

                section_qty_sum = df_section_data['商品数量'].sum()
                section_amt_sum = df_section_data['应结金额'].sum()

                qty_col_idx_detail = DETAIL_SHEET_COLUMNS_TMALL.index('商品数量') + 1
                cell_qty_detail = current_sheet_obj.cell(row=total_row_idx_for_section, column=qty_col_idx_detail, value=section_qty_sum)
                cell_qty_detail.font = bold_font; cell_qty_detail.number_format = '#,##0'

                amt_col_idx_detail = DETAIL_SHEET_COLUMNS_TMALL.index('应结金额') + 1
                cell_amt_detail = current_sheet_obj.cell(row=total_row_idx_for_section, column=amt_col_idx_detail, value=section_amt_sum)
                cell_amt_detail.font = bold_font; cell_amt_detail.number_format = '#,##0.00'

        write_section_to_sheet_detail(product_detail_sheet, detail_income_df_data, True, "收入总计")

        if not detail_expenditure_df_data.empty or True:
            product_detail_sheet.append([])
            write_section_to_sheet_detail(
                product_detail_sheet,
                detail_expenditure_df_data,
                not header_written_for_this_sheet,
                "支出总计"
            )

        if header_written_for_this_sheet:
            for current_col_idx_detail, current_column_title_detail in enumerate(DETAIL_SHEET_COLUMNS_TMALL, 1):
                column_letter_val_detail_sheet = get_column_letter(current_col_idx_detail)
                max_len_content = len(str(current_column_title_detail))
                for row_num_val_detail_sheet in range(1, product_detail_sheet.max_row + 1):
                    cell_val_obj_detail_sheet = product_detail_sheet.cell(row=row_num_val_detail_sheet, column=current_col_idx_detail).value
                    if cell_val_obj_detail_sheet is not None:
                        if isinstance(cell_val_obj_detail_sheet, (int, float)) and cell_val_obj_detail_sheet != 0:
                            if current_column_title_detail == '商品数量':
                                max_len_content = max(max_len_content, len(f"{cell_val_obj_detail_sheet:,}"))
                            elif current_column_title_detail in ['应结金额', '商品价格']:
                                 max_len_content = max(max_len_content, len(f"{cell_val_obj_detail_sheet:,.2f}"))
                            else:
                                max_len_content = max(max_len_content, len(str(cell_val_obj_detail_sheet)))
                        else:
                             max_len_content = max(max_len_content, len(str(cell_val_obj_detail_sheet)))

                adjusted_col_width = min(max(max_len_content + 4, 12), 60)
                if current_column_title_detail == "商品名称": adjusted_col_width = min(max(max_len_content + 4, 40), 70)
                elif current_column_title_detail == "商品属性": adjusted_col_width = min(max(max_len_content + 4, 30), 60)
                elif current_column_title_detail in ['订单创建时间', '订单付款时间', '发货时间']: adjusted_col_width = min(max(max_len_content + 4, 19), 25)
                elif current_column_title_detail in ['订单编号', '子订单编号', '商品编号', '物流单号']: adjusted_col_width = min(max(max_len_content +4, 22), 35)
                product_detail_sheet.column_dimensions[column_letter_val_detail_sheet].width = adjusted_col_width

    # --- 6. 确保 "销售总结" 工作表为第一个 ---
    if wb.sheetnames[0] != "销售总结" and "销售总结" in wb.sheetnames:
        summary_s_obj = wb["销售总结"]; wb._sheets.remove(summary_s_obj); wb._sheets.insert(0, summary_s_obj)

    # --- 7. 保存Excel工作簿 ---
    try:
        wb.save(output_file_path)
        return output_file_path
    except Exception as e_save_err:
        print(f"保存Excel文件 '{output_file_path}' 失败: {e_save_err}")
        alt_output_file_path_val = os.path.join(input_dir, f"{file_name_without_ext}_output_tmall_alt.xlsx")
        try:
            wb.save(alt_output_file_path_val)
            print(f"已尝试使用备用名称保存: {alt_output_file_path_val}")
            return alt_output_file_path_val
        except Exception as e_alt_save_err:
            print(f"使用备用名称保存也失败: {e_alt_save_err}")
            return None

# ---- 主程序入口 ----
if __name__ == "__main__":
    directory_path = r"F:\étude\Ecole\E4\E4stage\E4stageProjet\samples\TM"
    xlsx_filename = "ExportOrderList24220674019.xlsx" # 文件名

    input_file = os.path.join(directory_path, xlsx_filename)

    if not os.path.exists(input_file):
        print(f"错误: 输入文件未找到于 '{os.path.abspath(input_file)}'")
    elif not (input_file.lower().endswith(".xlsx") or input_file.lower().endswith(".xls")):
        print(f"错误: 输入文件 '{os.path.basename(input_file)}' 不是一个有效的Excel文件。")
    else:
        try:
            print(f"开始处理文件: {input_file}")
            output_file_result = process_tmall_sales_data(input_file)
            if output_file_result:
                print(f"处理完成。输出文件位于: {output_file_result}")
            else:
                print("处理过程中发生错误，未能生成输出文件。")
        except pd.errors.EmptyDataError:
            print(f"错误: 输入文件 '{os.path.basename(input_file)}' 为空或Pandas无法解析。")
        except KeyError as e_key:
            print(f"处理过程中发生KeyError错误: 列名 {e_key} 可能在输入文件中不存在，或与脚本配置不符。")
            print(f"请仔细核对脚本顶部的 TMALL_COL_... 常量是否与您Excel文件 '{xlsx_filename}' 的表头完全匹配。")
        except Exception as e_general:
            print(f"处理过程中发生未预料的错误: {e_general}")
            import traceback
            traceback.print_exc()