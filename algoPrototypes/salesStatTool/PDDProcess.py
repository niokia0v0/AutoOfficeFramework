import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import numpy as np

# Pandas 显示选项
pd.set_option('future.no_silent_downcasting', True)

# --- 配置项 ---
# 拼多多原始列名映射
PDD_COL_PRODUCT_NAME = '商品'
PDD_COL_ORDER_ID = '订单号'
PDD_COL_ORDER_STATUS = '订单状态'
PDD_COL_PRODUCT_TOTAL_PRICE = '商品总价(元)'
PDD_COL_STORE_DISCOUNT = '店铺优惠折扣(元)'
PDD_COL_PLATFORM_DISCOUNT = '平台优惠折扣(元)'
PDD_COL_USER_ACTUAL_PAYMENT = '用户实付金额(元)'
PDD_COL_MERCHANT_ACTUAL_RECEIPT = '商家实收金额(元)'
PDD_COL_QUANTITY = '商品数量(件)'
PDD_COL_SHIPPING_TIME = '发货时间'
PDD_COL_CONFIRM_RECEIPT_TIME = '确认收货时间'
PDD_COL_PRODUCT_ID = '商品id'
PDD_COL_PRODUCT_SPEC = '商品规格'
PDD_COL_AFTER_SALES_STATUS = '售后状态'
PDD_COL_LOGISTICS_NO = '快递单号'
PDD_COL_LOGISTICS_COMPANY = '快递公司'
PDD_COL_ORDER_TRANSACTION_TIME = '订单成交时间'

# 订单状态常量
STATUS_REFUND_SUCCESS = '退款成功'

# 输出到详情页的列定义
DETAIL_SHEET_COLUMNS_PDD = [
    '订单号', '订单状态', '售后状态', '商品ID', '商品名称', '商品规格', '商品数量(件)',
    '用户实付金额(元)', '商家实收金额(元)', '订单成交时间', '发货时间', '确认收货时间',
    '快递单号', '快递公司'
]

def process_pdd_sales_data(input_file_path):
    # --- 1. 初始化路径和读取数据 ---
    input_dir = os.path.dirname(input_file_path)
    base_name = os.path.basename(input_file_path)
    file_name_without_ext = os.path.splitext(base_name)[0]
    output_file_path = os.path.join(input_dir, f"{file_name_without_ext}_output_PDD.xlsx")

    try:
        df_original = pd.read_csv(input_file_path, dtype=str, encoding='utf-8-sig')
    except FileNotFoundError: return print(f"错误: 输入文件未找到于 '{os.path.abspath(input_file_path)}'")
    except Exception as e: return print(f"读取CSV文件 '{input_file_path}' 失败: {e}")

    if df_original.empty: return None

    # --- 2. 数据清洗和预处理 ---
    df_original.columns = df_original.columns.str.strip()
    for col in df_original.columns:
        if df_original[col].dtype == 'object':
            df_original[col] = df_original[col].astype(str).str.strip().replace(
                ['-', '--', '', 'None', 'nan', '#NULL!', None, 'null', '\t'], np.nan, regex=False)

    # **修正点**: 修复了变量名的拼写错误
    numeric_cols = [
        PDD_COL_QUANTITY, PDD_COL_USER_ACTUAL_PAYMENT, PDD_COL_MERCHANT_ACTUAL_RECEIPT,
        PDD_COL_PRODUCT_TOTAL_PRICE, PDD_COL_STORE_DISCOUNT, PDD_COL_PLATFORM_DISCOUNT,
    ]
    for col in numeric_cols:
        if col in df_original.columns:
            df_original[col] = pd.to_numeric(df_original[col], errors='coerce').fillna(0)

    # --- 3. 核心逻辑：数据筛选与分组 ---
    df_no_cancel = df_original[~df_original[PDD_COL_ORDER_STATUS].str.contains('取消', na=False)].copy()
    df_processed = df_no_cancel[df_no_cancel[PDD_COL_PRODUCT_ID].notna()].copy()
    df_processed[PDD_COL_PRODUCT_ID] = df_processed[PDD_COL_PRODUCT_ID].astype(str)

    if df_processed.empty: return None

    # --- 4. 按商品ID汇总信息 ---
    product_data_map = {}
    
    def format_df_for_detail(df_source, p_id, p_name, is_refund=False):
        if df_source.empty: return pd.DataFrame(columns=DETAIL_SHEET_COLUMNS_PDD)
        df_target = pd.DataFrame({
            '订单号': df_source.get(PDD_COL_ORDER_ID), '订单状态': df_source.get(PDD_COL_ORDER_STATUS),
            '售后状态': df_source.get(PDD_COL_AFTER_SALES_STATUS), '商品ID': p_id, '商品名称': p_name,
            '商品规格': df_source.get(PDD_COL_PRODUCT_SPEC), '商品数量(件)': df_source.get(PDD_COL_QUANTITY),
            '用户实付金额(元)': df_source.get(PDD_COL_USER_ACTUAL_PAYMENT), '商家实收金额(元)': 0,
            '订单成交时间': df_source.get(PDD_COL_ORDER_TRANSACTION_TIME), '发货时间': df_source.get(PDD_COL_SHIPPING_TIME),
            '确认收货时间': df_source.get(PDD_COL_CONFIRM_RECEIPT_TIME), '快递单号': df_source.get(PDD_COL_LOGISTICS_NO),
            '快递公司': df_source.get(PDD_COL_LOGISTICS_COMPANY)
        })
        if not is_refund:
            df_target['商家实收金额(元)'] = df_source[PDD_COL_PRODUCT_TOTAL_PRICE] - df_source[PDD_COL_STORE_DISCOUNT]
        else:
            df_target['商家实收金额(元)'] = -(df_source[PDD_COL_USER_ACTUAL_PAYMENT] + df_source[PDD_COL_PLATFORM_DISCOUNT])
            df_target['用户实付金额(元)'] = -df_source[PDD_COL_USER_ACTUAL_PAYMENT]
        return df_target[DETAIL_SHEET_COLUMNS_PDD]

    for product_id, group_df in df_processed.groupby(PDD_COL_PRODUCT_ID):
        product_name = group_df[PDD_COL_PRODUCT_NAME].iloc[0] if not group_df[PDD_COL_PRODUCT_NAME].empty else "未知商品"

        all_refunds_df = group_df[group_df[PDD_COL_AFTER_SALES_STATUS].str.contains(STATUS_REFUND_SUCCESS, na=False)].copy()
        regular_refund_df = all_refunds_df[all_refunds_df[PDD_COL_ORDER_STATUS].str.contains('未发货', na=False)].copy()
        shipped_refund_df = all_refunds_df[all_refunds_df[PDD_COL_ORDER_STATUS].str.contains('已发货', na=False)].copy()

        product_data_map[product_id] = {
            'name': product_name,
            'income_df': format_df_for_detail(group_df, product_id, product_name),
            'unshipped_refund_df': format_df_for_detail(regular_refund_df, product_id, product_name, is_refund=True),
            'shipped_refund_df': format_df_for_detail(shipped_refund_df, product_id, product_name, is_refund=True)
        }

    # --- 5. 创建Excel工作簿并写入销售总结页 ---
    wb = Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "销售总结"
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    
    current_row = 1
    
    # 收入汇总
    summary_sheet.cell(row=current_row, column=1, value="各商品收入汇总 (所有未取消订单)").font = bold_font
    current_row += 1
    summary_sheet.append(["商品ID", "商品名称", "总销售数量", "用户实付总额(参考)", "总销售额"])
    for cell in summary_sheet[current_row]: cell.font = bold_font; cell.alignment = center_align
    current_row += 1
    
    sorted_product_ids = sorted(product_data_map.keys())
    grand_total_income_qty, grand_total_income_user_pay, grand_total_income_receipt = 0, 0, 0
    for prod_id in sorted_product_ids:
        item = product_data_map[prod_id]
        df = item['income_df']
        qty, user_pay, receipt = df['商品数量(件)'].sum(), df['用户实付金额(元)'].sum(), df['商家实收金额(元)'].sum()
        summary_sheet.append([prod_id, item['name'], qty, user_pay, receipt])
        grand_total_income_qty += qty
        grand_total_income_user_pay += user_pay
        grand_total_income_receipt += receipt
        current_row += 1
    summary_sheet.append(["总计销售", "", grand_total_income_qty, grand_total_income_user_pay, grand_total_income_receipt])
    for cell in summary_sheet[current_row]: cell.font = bold_font
    current_row += 2

    # 未发货退款汇总
    summary_sheet.cell(row=current_row, column=1, value="各商品支出汇总 (未发货退款)").font = bold_font
    current_row += 1
    summary_sheet.append(["商品ID", "商品名称", "退款数量", "用户实付总额(退款)", "总退款额"])
    for cell in summary_sheet[current_row]: cell.font = bold_font; cell.alignment = center_align
    current_row += 1
    
    grand_total_unshipped_qty, grand_total_unshipped_user_pay, grand_total_unshipped_receipt = 0, 0, 0
    for prod_id in sorted_product_ids:
        item = product_data_map[prod_id]
        df = item['unshipped_refund_df']
        if not df.empty:
            qty, user_pay, receipt = df['商品数量(件)'].sum(), df['用户实付金额(元)'].sum(), df['商家实收金额(元)'].sum()
            summary_sheet.append([prod_id, item['name'], qty, user_pay, receipt])
            grand_total_unshipped_qty += qty
            grand_total_unshipped_user_pay += user_pay
            grand_total_unshipped_receipt += receipt
            current_row += 1
    summary_sheet.append(["未发货退款总计", "", grand_total_unshipped_qty, grand_total_unshipped_user_pay, grand_total_unshipped_receipt])
    for cell in summary_sheet[current_row]: cell.font = bold_font
    current_row += 2

    # 已发货退款汇总
    summary_sheet.cell(row=current_row, column=1, value="各商品支出汇总 (已发货退款)").font = bold_font
    current_row += 1
    summary_sheet.append(["商品ID", "商品名称", "退款数量", "用户实付(退款)", "退款额"])
    for cell in summary_sheet[current_row]: cell.font = bold_font; cell.alignment = center_align
    current_row += 1

    grand_total_shipped_qty, grand_total_shipped_user_pay, grand_total_shipped_receipt = 0, 0, 0
    for prod_id in sorted_product_ids:
        item = product_data_map[prod_id]
        df = item['shipped_refund_df']
        if not df.empty:
            qty, user_pay, receipt = df['商品数量(件)'].sum(), df['用户实付金额(元)'].sum(), df['商家实收金额(元)'].sum()
            summary_sheet.append([prod_id, item['name'], qty, user_pay, receipt])
            grand_total_shipped_qty += qty
            grand_total_shipped_user_pay += user_pay
            grand_total_shipped_receipt += receipt
            current_row += 1
    summary_sheet.append(["已发货退款总计", "", grand_total_shipped_qty, grand_total_shipped_user_pay, grand_total_shipped_receipt])
    for cell in summary_sheet[current_row]: cell.font = bold_font
    current_row += 2

    # 两种净总计
    net_qty_view1 = grand_total_income_qty - grand_total_unshipped_qty
    net_user_pay_view1 = grand_total_income_user_pay + grand_total_unshipped_user_pay
    net_receipt_view1 = grand_total_income_receipt + grand_total_unshipped_receipt
    
    summary_sheet.cell(row=current_row, column=1, value="净总计(已发货退款订单按售出计算)").font = bold_font
    summary_sheet.cell(row=current_row, column=3, value=net_qty_view1).font = bold_font
    summary_sheet.cell(row=current_row, column=4, value=net_user_pay_view1).font = bold_font
    summary_sheet.cell(row=current_row, column=5, value=net_receipt_view1).font = bold_font
    current_row += 1

    net_qty_view2 = grand_total_income_qty - grand_total_unshipped_qty - grand_total_shipped_qty
    net_user_pay_view2 = grand_total_income_user_pay + grand_total_unshipped_user_pay + grand_total_shipped_user_pay
    net_receipt_view2 = grand_total_income_receipt + grand_total_unshipped_receipt + grand_total_shipped_receipt
    
    summary_sheet.cell(row=current_row, column=1, value="净总计(已发货退款订单按退款计算)").font = bold_font
    summary_sheet.cell(row=current_row, column=3, value=net_qty_view2).font = bold_font
    summary_sheet.cell(row=current_row, column=4, value=net_user_pay_view2).font = bold_font
    summary_sheet.cell(row=current_row, column=5, value=net_receipt_view2).font = bold_font
    
    # 格式化总结页
    for col_letter, width in [('A', 35), ('B', 60), ('C', 18), ('D', 22), ('E', 22)]:
        summary_sheet.column_dimensions[col_letter].width = width
    for row in summary_sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.column in [3]: cell.number_format = '#,##0'
                if cell.column in [4, 5]: cell.number_format = '#,##0.00'

    # --- 6. 为每个商品创建并写入详情页 ---
    for prod_id in sorted_product_ids:
        item = product_data_map[prod_id]
        sheet_name = re.sub(r'[\\/\*\[\]\:?]', '_', f"{prod_id}_{item['name']}")[:31]
        try: ws = wb.create_sheet(sheet_name)
        except: ws = wb.create_sheet(f"{prod_id}_detail")

        def write_section(sheet, df, title, total_title):
            if df.empty: return
            sheet.append([title])
            sheet[sheet.max_row][0].font = bold_font
            sheet.append(DETAIL_SHEET_COLUMNS_PDD)
            for cell in sheet[sheet.max_row]: cell.font = bold_font; cell.alignment = center_align
            for _, row_data in df.iterrows(): sheet.append(row_data.tolist())
            total_row_data = [total_title, "", "", "", "", "", df['商品数量(件)'].sum(), df['用户实付金额(元)'].sum(), df['商家实收金额(元)'].sum()]
            sheet.append(total_row_data)
            for cell in sheet[sheet.max_row]: cell.font = bold_font
            sheet.append([])

        write_section(ws, item['income_df'], "收入明细 (所有未取消订单)", "收入总计")
        write_section(ws, item['unshipped_refund_df'], "支出明细 (未发货退款)", "未发货退款总计")
        write_section(ws, item['shipped_refund_df'], "支出明细 (已发货退款)", "已发货退款总计")

        for col_idx, width in enumerate([25, 20, 15, 25, 60, 22, 15, 18, 18, 20, 20, 20, 25, 15], 1):
             ws.column_dimensions[get_column_letter(col_idx)].width = width

    if wb.sheetnames[0] != "销售总结": wb.move_sheet("销售总结", -len(wb.sheetnames)+1)
    
    try:
        wb.save(output_file_path)
        return output_file_path
    except Exception as e:
        print(f"保存Excel文件失败: {e}")
        return None

# ---- 主程序入口 ----
if __name__ == "__main__":
    directory_path = r"F:\étude\Ecole\E4\E4stage\E4stageProjet\samples\PDD"
    csv_filename = "931601647orders_export2025-05-29-17-15-09.csv"
    input_file = os.path.join(directory_path, csv_filename)

    if not os.path.exists(input_file):
        print(f"错误: 输入文件未找到于 '{os.path.abspath(input_file)}'")
    else:
        try:
            print(f"开始处理文件: {input_file}")
            output_file_result = process_pdd_sales_data(input_file)
            if output_file_result:
                print(f"处理完成。输出文件位于: {output_file_result}")
            else:
                print("处理过程中发生错误，未能生成输出文件。")
        except Exception as e:
            print(f"处理过程中发生未预料的错误: {e}")
            import traceback
            traceback.print_exc()