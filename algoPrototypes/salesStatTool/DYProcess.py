import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import numpy as np

# Pandas 显示选项
pd.set_option('future.no_silent_downcasting', True)

# --- 配置项 ---
# 抖店原始列名映射
DY_COL_MAIN_ORDER_ID = '主订单编号'
DY_COL_PRODUCT_NAME_DESC = '选购商品'
DY_COL_SKU = '货号'
DY_COL_PRODUCT_ID = '商品ID'
DY_COL_QUANTITY = '商品数量'
DY_COL_UNIT_PRICE = '商品金额' # 注意：这是单价
DY_COL_ORDER_SUBMIT_TIME = '订单提交时间'
DY_COL_PAY_COMPLETE_TIME = '支付完成时间'
DY_COL_ORDER_COMPLETE_TIME = '订单完成时间'
DY_COL_ORDER_STATUS = '订单状态'
DY_COL_CANCEL_REASON = '取消原因'
DY_COL_AFTER_SALES_STATUS = '售后状态'
DY_COL_ORDER_TYPE = '订单类型'

# 订单状态常量
STATUS_COMPLETED = '已完成'
STATUS_CLOSED = '已关闭' # 可能还有其他非完成状态

# 输出到详情页的列定义
DETAIL_SHEET_COLUMNS_DY = [
    '订单编号',
    '订单状态',
    '售后状态',
    '取消原因',
    '商品编号',
    '商品名称',
    '商品单价',
    '商品数量',
    '应付款',
    '订单提交时间',
    '支付完成时间',
    '订单完成时间',
]

def process_douyin_sales_data(input_file_path):
    """
    处理抖店订单导出的CSV文件，生成包含销售总结和各商品销售明细的Excel文件。
    总结页和详情页均按“收入”和“支出”逻辑进行分区。
    收入部分包含全部订单，支出部分只包含状态不为“已完成”的订单。
    详情页增加“应付款”列。支出项金额显示为负数。
    """
    # --- 1. 初始化路径和读取数据 ---
    input_dir = os.path.dirname(input_file_path)
    base_name = os.path.basename(input_file_path)
    file_name_without_ext = os.path.splitext(base_name)[0]
    output_file_path = os.path.join(input_dir, f"{file_name_without_ext}_output_DY.xlsx")

    try:
        df_original = pd.read_csv(input_file_path, dtype=str, keep_default_na=True)
        df_original.columns = [col.strip().replace('"', '') for col in df_original.columns]
    except FileNotFoundError:
        print(f"错误: 输入文件未找到于 '{os.path.abspath(input_file_path)}'")
        return None
    except Exception as e:
        print(f"读取CSV文件 '{input_file_path}' 失败: {e}")
        return None

    if df_original.empty:
        print(f"错误: 输入文件 '{base_name}' 为空或未能正确读取数据。")
        return None

    # --- 2. 数据清洗和预处理 ---
    df_original.columns = df_original.columns.str.strip()
    for col in df_original.columns:
        if df_original[col].dtype == 'object':
            df_original[col] = df_original[col].astype(str).str.strip().str.replace('"', '').str.replace('\t', '')
            df_original[col] = df_original[col].replace(
                ['-', '--', '', 'None', 'nan', '#NULL!', None, 'null'], np.nan, regex=False
            )

    critical_logic_cols_check = {
        DY_COL_PRODUCT_ID: "商品ID",
        DY_COL_ORDER_STATUS: "订单状态",
        DY_COL_QUANTITY: "商品数量",
        DY_COL_UNIT_PRICE: "商品金额",
    }
    for col_const, col_display_name in critical_logic_cols_check.items():
        if col_const not in df_original.columns:
            print(f"错误: 核心逻辑所需列 '{col_display_name}' (脚本常量名: {col_const}) "
                  f"在输入文件 '{base_name}' 的表头中未找到。脚本无法继续。")
            return None

    numeric_cols_to_convert_map = {
        DY_COL_QUANTITY: 0,
        DY_COL_UNIT_PRICE: 0.0,
    }
    for col, default_fill_value in numeric_cols_to_convert_map.items():
        if col in df_original.columns:
            df_original[col] = pd.to_numeric(df_original[col], errors='coerce').fillna(default_fill_value)
        else:
            print(f"警告: 列 '{col}' 在输入文件中未找到，将创建并填充默认值 {default_fill_value}。")
            df_original[col] = default_fill_value

    df_original[DY_COL_PRODUCT_ID] = df_original[DY_COL_PRODUCT_ID].astype(str).replace('nan', np.nan)
    df_processed = df_original[df_original[DY_COL_PRODUCT_ID].notna()].copy()

    if df_processed.empty:
        print(f"数据中没有找到包含有效商品ID ('{DY_COL_PRODUCT_ID}') 的行。无法生成报告。")
        return None

    df_processed['应付款'] = df_processed[DY_COL_UNIT_PRICE] * df_processed[DY_COL_QUANTITY]

    # --- 3. 按商品ID汇总信息 ---
    product_data_map = {}
    for product_id_value, group_df in df_processed.groupby(DY_COL_PRODUCT_ID):
        product_id_str_key = str(product_id_value)
        product_name_series_data = group_df[DY_COL_PRODUCT_NAME_DESC].dropna() if DY_COL_PRODUCT_NAME_DESC in group_df else pd.Series([])
        product_name_str = product_name_series_data.iloc[0] if not product_name_series_data.empty else "未知商品"

        income_orders_group = group_df.copy()
        income_total_quantity_per_product = income_orders_group[DY_COL_QUANTITY].sum()
        income_total_payable_amount_per_product = income_orders_group['应付款'].sum()

        expenditure_orders_group_orig = group_df[group_df[DY_COL_ORDER_STATUS] != STATUS_COMPLETED].copy()
        # 复制一份用于修改金额为负数，而不影响原始的 group_df 或 df_processed
        expenditure_orders_group_for_detail = expenditure_orders_group_orig.copy()
        if not expenditure_orders_group_for_detail.empty:
            expenditure_orders_group_for_detail['应付款'] = expenditure_orders_group_for_detail['应付款'] * -1

        expenditure_total_quantity_per_product = expenditure_orders_group_orig[DY_COL_QUANTITY].sum()
        # 支出总金额直接从修改后的 dataframe 中汇总，或者原始汇总后乘以-1
        expenditure_total_payable_amount_per_product = expenditure_orders_group_for_detail['应付款'].sum() # 这已经是负数了
        # 或者： expenditure_orders_group_orig['应付款'].sum() * -1


        def format_df_for_detail_sheet_dy(source_df_input, prod_id_for_detail, prod_name_for_detail):
            if source_df_input.empty:
                return pd.DataFrame(columns=DETAIL_SHEET_COLUMNS_DY)

            detail_df_formatted = pd.DataFrame()
            detail_df_formatted['订单编号'] = source_df_input.get(DY_COL_MAIN_ORDER_ID)
            detail_df_formatted['订单状态'] = source_df_input.get(DY_COL_ORDER_STATUS)
            detail_df_formatted['售后状态'] = source_df_input.get(DY_COL_AFTER_SALES_STATUS)
            detail_df_formatted['取消原因'] = source_df_input.get(DY_COL_CANCEL_REASON)
            detail_df_formatted['商品编号'] = prod_id_for_detail
            detail_df_formatted['商品名称'] = prod_name_for_detail
            detail_df_formatted['商品单价'] = source_df_input.get(DY_COL_UNIT_PRICE)
            detail_df_formatted['商品数量'] = source_df_input.get(DY_COL_QUANTITY)
            detail_df_formatted['应付款'] = source_df_input.get('应付款') # 这里的应付款已经是处理过的（支出部分为负）
            detail_df_formatted['订单提交时间'] = source_df_input.get(DY_COL_ORDER_SUBMIT_TIME)
            detail_df_formatted['支付完成时间'] = source_df_input.get(DY_COL_PAY_COMPLETE_TIME)
            detail_df_formatted['订单完成时间'] = source_df_input.get(DY_COL_ORDER_COMPLETE_TIME)
            
            final_detail_df = pd.DataFrame(columns=DETAIL_SHEET_COLUMNS_DY)
            for col_name in DETAIL_SHEET_COLUMNS_DY:
                if col_name in detail_df_formatted.columns:
                    final_detail_df[col_name] = detail_df_formatted[col_name]
                else:
                    if col_name in ['商品单价', '商品数量', '应付款']:
                        final_detail_df[col_name] = 0 if col_name == '商品数量' else 0.0
                    else:
                        final_detail_df[col_name] = ''
            return final_detail_df

        detail_income_section_df = format_df_for_detail_sheet_dy(
            income_orders_group, product_id_str_key, product_name_str # 收入部分使用原始应付款
        )
        detail_expenditure_section_df = format_df_for_detail_sheet_dy(
            expenditure_orders_group_for_detail, product_id_str_key, product_name_str # 支出部分使用金额已为负的DF
        )

        product_data_map[product_id_str_key] = {
            'name': product_name_str,
            'income_total_quantity': income_total_quantity_per_product,
            'income_total_payable_amount': income_total_payable_amount_per_product, # 正数
            'expenditure_total_quantity': expenditure_total_quantity_per_product,
            'expenditure_total_payable_amount': expenditure_total_payable_amount_per_product, # 负数
            'detail_income_df': detail_income_section_df, # 应付款为正
            'detail_expenditure_df': detail_expenditure_section_df, # 应付款为负
        }

    # --- 4. 创建Excel工作簿并写入销售总结页 ---
    wb = Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "销售总结"

    summary_income_header = ["商品编号", "商品名称", "总销售数量 (全部订单)", "总应付金额 (全部订单)"]
    summary_expenditure_header = ["商品编号", "商品名称", "未完成订单商品数量", "未完成订单应付金额 (支出)"]
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')

    current_row_summary = 1

    summary_sheet.cell(row=current_row_summary, column=1, value="各商品收入汇总").font = bold_font
    current_row_summary += 1
    summary_sheet.append(summary_income_header)
    for cell in summary_sheet[current_row_summary]: cell.font = bold_font; cell.alignment = center_alignment
    current_row_summary += 1

    grand_total_income_qty = 0
    grand_total_income_amt = 0.0 # 明确为浮点数
    sorted_product_ids = sorted(product_data_map.keys())

    for prod_id in sorted_product_ids:
        item_data = product_data_map[prod_id]
        summary_sheet.append([
            prod_id, item_data['name'],
            item_data['income_total_quantity'], item_data['income_total_payable_amount']
        ])
        grand_total_income_qty += item_data['income_total_quantity']
        grand_total_income_amt += item_data['income_total_payable_amount']
        summary_sheet.cell(row=current_row_summary, column=3).number_format = '#,##0'
        summary_sheet.cell(row=current_row_summary, column=4).number_format = '#,##0.00'
        current_row_summary += 1

    summary_sheet.cell(row=current_row_summary, column=1, value="总计收入").font = bold_font
    summary_sheet.cell(row=current_row_summary, column=3, value=grand_total_income_qty).font = bold_font
    summary_sheet.cell(row=current_row_summary, column=3).number_format = '#,##0'
    summary_sheet.cell(row=current_row_summary, column=4, value=grand_total_income_amt).font = bold_font
    summary_sheet.cell(row=current_row_summary, column=4).number_format = '#,##0.00'
    current_row_summary += 1
    summary_sheet.append([])
    current_row_summary += 1

    summary_sheet.cell(row=current_row_summary, column=1, value="各商品支出汇总 (未完成订单)").font = bold_font
    current_row_summary += 1
    summary_sheet.append(summary_expenditure_header)
    for cell in summary_sheet[current_row_summary]: cell.font = bold_font; cell.alignment = center_alignment
    current_row_summary += 1

    grand_total_expenditure_qty = 0
    grand_total_expenditure_amt = 0.0 # 明确为浮点数

    for prod_id in sorted_product_ids:
        item_data = product_data_map[prod_id]
        if item_data['expenditure_total_quantity'] > 0 or item_data['expenditure_total_payable_amount'] != 0:
            summary_sheet.append([
                prod_id, item_data['name'],
                item_data['expenditure_total_quantity'], item_data['expenditure_total_payable_amount'] # 已经是负数
            ])
            grand_total_expenditure_qty += item_data['expenditure_total_quantity']
            grand_total_expenditure_amt += item_data['expenditure_total_payable_amount'] # 累加负数
            summary_sheet.cell(row=current_row_summary, column=3).number_format = '#,##0'
            summary_sheet.cell(row=current_row_summary, column=4).number_format = '#,##0.00' # 负数也会正确显示
            current_row_summary += 1

    summary_sheet.cell(row=current_row_summary, column=1, value="总计支出 (未完成订单)").font = bold_font
    summary_sheet.cell(row=current_row_summary, column=3, value=grand_total_expenditure_qty).font = bold_font
    summary_sheet.cell(row=current_row_summary, column=3).number_format = '#,##0'
    summary_sheet.cell(row=current_row_summary, column=4, value=grand_total_expenditure_amt).font = bold_font # 已经是负数总和
    summary_sheet.cell(row=current_row_summary, column=4).number_format = '#,##0.00'
    current_row_summary += 1
    summary_sheet.append([])
    current_row_summary += 1

    net_total_quantity = grand_total_income_qty - grand_total_expenditure_qty
    net_total_amount = grand_total_income_amt + grand_total_expenditure_amt # 因为支出总额已是负数

    summary_sheet.cell(row=current_row_summary, column=1, value="净总计").font = bold_font
    summary_sheet.cell(row=current_row_summary, column=3, value=net_total_quantity).font = bold_font
    summary_sheet.cell(row=current_row_summary, column=3).number_format = '#,##0'
    summary_sheet.cell(row=current_row_summary, column=4, value=net_total_amount).font = bold_font
    summary_sheet.cell(row=current_row_summary, column=4).number_format = '#,##0.00'

    summary_sheet.column_dimensions['A'].width = 25
    summary_sheet.column_dimensions['B'].width = 70
    summary_sheet.column_dimensions['C'].width = 20
    summary_sheet.column_dimensions['D'].width = 25 # 适应 "未完成订单应付金额 (支出)"

    # --- 5. 为每个商品创建并写入详情页 ---
    for product_id_str_key in sorted_product_ids:
        product_info_item = product_data_map[product_id_str_key]
        detail_income_df_data = product_info_item['detail_income_df']       # 应付款为正
        detail_expenditure_df_data = product_info_item['detail_expenditure_df'] # 应付款为负

        if detail_income_df_data.empty and detail_expenditure_df_data.empty:
            continue

        clean_product_name_str = re.sub(r'[\\/\*\[\]\:?]', '_', str(product_info_item['name']))
        base_sheet_name = f"{product_id_str_key}_{clean_product_name_str}"
        if len(base_sheet_name) > 31:
            max_name_len = 31 - len(product_id_str_key) - 1
            if max_name_len < 3:
                sheet_name_final = base_sheet_name[:31]
            else:
                sheet_name_final = f"{product_id_str_key}_{clean_product_name_str[:max_name_len]}"
        else:
            sheet_name_final = base_sheet_name
        
        try:
            product_detail_sheet = wb.create_sheet(sheet_name_final)
        except Exception:
            product_detail_sheet = wb.create_sheet(base_sheet_name[:28] + "...")

        header_written_for_this_sheet = False

        def write_section_to_sheet_detail_dy(current_sheet_obj, df_section_data,
                                             is_header_needed_flag, section_title_str=None):
            nonlocal header_written_for_this_sheet
            if df_section_data.empty and not section_title_str:
                return

            df_to_write_to_excel = df_section_data.copy()
            df_to_write_to_excel = df_to_write_to_excel.reindex(columns=DETAIL_SHEET_COLUMNS_DY, fill_value='')

            for col_name_fmt in df_to_write_to_excel.columns:
                if col_name_fmt in ['商品单价', '商品数量', '应付款']:
                    df_to_write_to_excel[col_name_fmt] = pd.to_numeric(df_to_write_to_excel[col_name_fmt], errors='coerce').fillna(0)
                else:
                    df_to_write_to_excel[col_name_fmt] = df_to_write_to_excel[col_name_fmt].fillna('').astype(str)

            if is_header_needed_flag and not header_written_for_this_sheet:
                current_sheet_obj.append(DETAIL_SHEET_COLUMNS_DY)
                for cell_header_obj in current_sheet_obj[current_sheet_obj.max_row]:
                    cell_header_obj.font = bold_font
                    cell_header_obj.alignment = center_alignment
                header_written_for_this_sheet = True
            
            if not df_section_data.empty:
                for _, row_data_item in df_to_write_to_excel.iterrows():
                    current_sheet_obj.append(row_data_item.tolist())

            if section_title_str:
                total_row_idx_for_section = current_sheet_obj.max_row + 1
                current_sheet_obj.cell(row=total_row_idx_for_section, column=1, value=section_title_str).font = bold_font

                # 这里的sum是基于已经处理过金额符号的df_section_data
                section_qty_sum = df_section_data['商品数量'].sum() if '商品数量' in df_section_data and not df_section_data.empty else 0
                section_amt_sum = df_section_data['应付款'].sum() if '应付款' in df_section_data and not df_section_data.empty else 0.0

                if '商品数量' in DETAIL_SHEET_COLUMNS_DY:
                    qty_col_idx_detail = DETAIL_SHEET_COLUMNS_DY.index('商品数量') + 1
                    cell_qty_detail = current_sheet_obj.cell(row=total_row_idx_for_section, column=qty_col_idx_detail, value=section_qty_sum)
                    cell_qty_detail.font = bold_font; cell_qty_detail.number_format = '#,##0'

                if '应付款' in DETAIL_SHEET_COLUMNS_DY:
                    amt_col_idx_detail = DETAIL_SHEET_COLUMNS_DY.index('应付款') + 1
                    cell_amt_detail = current_sheet_obj.cell(row=total_row_idx_for_section, column=amt_col_idx_detail, value=section_amt_sum) # 金额已经是正确的符号
                    cell_amt_detail.font = bold_font; cell_amt_detail.number_format = '#,##0.00'
        
        write_section_to_sheet_detail_dy(product_detail_sheet, detail_income_df_data, True, "收入总计")
        product_detail_sheet.append([])
        write_section_to_sheet_detail_dy(
            product_detail_sheet,
            detail_expenditure_df_data,
            not header_written_for_this_sheet,
            "支出总计 (未完成订单)"
        )

        if header_written_for_this_sheet:
            for current_col_idx_detail, current_column_title_detail in enumerate(DETAIL_SHEET_COLUMNS_DY, 1):
                column_letter_val_detail_sheet = get_column_letter(current_col_idx_detail)
                max_len_content = len(str(current_column_title_detail))
                start_row_for_data = 2 if product_detail_sheet.max_row > 1 and header_written_for_this_sheet else 1
                for row_num_val_detail_sheet in range(start_row_for_data, product_detail_sheet.max_row + 1):
                    cell_val_obj_detail_sheet = product_detail_sheet.cell(row=row_num_val_detail_sheet, column=current_col_idx_detail).value
                    if cell_val_obj_detail_sheet is not None:
                        current_len = 0
                        if isinstance(cell_val_obj_detail_sheet, (int, float)):
                            if current_column_title_detail == '商品数量': current_len = len(f"{cell_val_obj_detail_sheet:,}")
                            elif current_column_title_detail in ['应付款', '商品单价']: current_len = len(f"{cell_val_obj_detail_sheet:,.2f}")
                            else: current_len = len(str(cell_val_obj_detail_sheet))
                        else: current_len = len(str(cell_val_obj_detail_sheet))
                        max_len_content = max(max_len_content, current_len)
                
                adjusted_col_width = min(max(max_len_content + 4, 12), 70)
                if current_column_title_detail == "商品名称": adjusted_col_width = min(max(max_len_content + 4, 40), 80)
                elif current_column_title_detail in ['订单提交时间', '支付完成时间', '订单完成时间']: adjusted_col_width = min(max(max_len_content + 4, 19), 25)
                elif current_column_title_detail in ['订单编号', '商品编号']: adjusted_col_width = min(max(max_len_content +4, 22), 35)
                product_detail_sheet.column_dimensions[column_letter_val_detail_sheet].width = adjusted_col_width

    # --- 6. 确保 "销售总结" 工作表为第一个 ---
    if wb.sheetnames[0] != "销售总结" and "销售总结" in wb.sheetnames:
        summary_s_obj = wb["销售总结"]; wb._sheets.remove(summary_s_obj); wb._sheets.insert(0, summary_s_obj)

    # --- 7. 保存Excel工作簿 ---
    try:
        wb.save(output_file_path)
        return output_file_path
    except Exception as e_save_err:
        print(f"保存Excel文件 '{output_file_path}' 失败: {e_save_err}")
        alt_output_file_path_val = os.path.join(input_dir, f"{file_name_without_ext}_output_dy_alt.xlsx")
        try:
            wb.save(alt_output_file_path_val)
            print(f"已尝试使用备用名称保存: {alt_output_file_path_val}")
            return alt_output_file_path_val
        except Exception as e_alt_save_err:
            print(f"使用备用名称保存也失败: {e_alt_save_err}")
            return None

# ---- 主程序入口 ----
if __name__ == "__main__":
    directory_path = r"F:\étude\Ecole\E4\E4stage\E4stageProjet\samples\DY"
    csv_filename = "抖店202504月.csv"
    input_file = os.path.join(directory_path, csv_filename)

    if not os.path.exists(input_file):
        print(f"错误: 输入文件未找到于 '{os.path.abspath(input_file)}'")
    elif not input_file.lower().endswith(".csv"):
        print(f"错误: 输入文件 '{os.path.basename(input_file)}' 不是一个有效的CSV文件。")
    else:
        try:
            print(f"开始处理文件: {input_file}")
            output_file_result = process_douyin_sales_data(input_file)
            if output_file_result:
                print(f"处理完成。输出文件位于: {output_file_result}")
            else:
                print("处理过程中发生错误，未能生成输出文件。")
        except pd.errors.EmptyDataError:
            print(f"错误: 输入文件 '{os.path.basename(input_file)}' 为空或Pandas无法解析。")
        except KeyError as e_key:
            print(f"处理过程中发生KeyError错误: 列名 {e_key} 可能在输入文件中不存在，或与脚本配置不符。")
            print(f"请仔细核对脚本顶部的 DY_COL_... 常量是否与您CSV文件 '{csv_filename}' 的表头完全匹配。")
        except Exception as e_general:
            print(f"处理过程中发生未预料的错误: {e_general}")
            import traceback
            traceback.print_exc()