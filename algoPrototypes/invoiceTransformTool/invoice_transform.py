import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from datetime import datetime
from dateutil.relativedelta import relativedelta
import os

# --- 1. 配置区域 ---
INPUT_FOLDER = r'C:\Users\LENOVO\Desktop'
OUTPUT_FOLDER = r'C:\Users\LENOVO\Desktop'
INPUT_FILENAME = 'desens_发票记录格式.xlsx'
OUTPUT_FILENAME = 'out_desens_发票记录格式.xlsx'

# --- 2. 辅助函数 ---
def calculate_expiry_date(batch_str):
    """
    根据6位批号字符串（YYMMDD格式）计算有效期。
    有效期规则为：生产日期 + 2年 - 1天。
    返回一个 "YYYYMMDD" 格式的字符串。
    """
    # 检查批号是否为6位数字字符串，否则返回空
    if not isinstance(batch_str, str) or len(batch_str) != 6 or not batch_str.isdigit():
        return ""
    try:
        # 将 "YYMMDD" 转换为 "20YYMMDD" 并解析为日期对象
        start_date = datetime.strptime(f"20{batch_str}", "%Y%m%d")
        # 计算有效期
        expiry_date = start_date + relativedelta(years=2) - relativedelta(days=1)
        # 将计算出的日期格式化为 "YYYYMMDD" 字符串并返回
        return expiry_date.strftime("%Y%m%d")
    except ValueError:
        # 如果日期格式无效，则捕获错误并返回空
        return ""

# --- 3. 主处理函数 ---
def process_sales_data():
    """
    主函数，执行从读取、处理、排序到生成格式化Excel文件的所有步骤。
    """
    # 构造完整的文件路径
    input_path = os.path.join(INPUT_FOLDER, INPUT_FILENAME)
    output_path = os.path.join(OUTPUT_FOLDER, OUTPUT_FILENAME)
    print(f"开始处理文件: {input_path}")

    # --- 步骤 1: 加载源数据 ---
    try:
        # 使用pandas读取Excel，将第一行作为表头
        df = pd.read_excel(input_path, header=0)
        # 同时使用openpyxl加载工作簿，以便后续读取单元格格式（如颜色）
        workbook_ole = openpyxl.load_workbook(input_path)
        sheet_ole = workbook_ole.active
    except FileNotFoundError:
        print(f"错误：输入文件未找到 -> {input_path}")
        return
        
    # --- 步骤 2: 数据预处理和准备 ---
    # 删除'单号'列为空的行，这些行是无效数据
    df.dropna(subset=['单号'], inplace=True, how='all')
    df.reset_index(drop=True, inplace=True)

    # 在排序前，将字体颜色信息作为一列添加到DataFrame中，使其与数据行绑定
    font_colors = []
    for index, row in df.iterrows():
        # 定位到原始Excel文件中对应的单元格
        customer_cell = sheet_ole.cell(row=index + 2, column=df.columns.get_loc('客户名称') + 1)
        source_font_color = None
        # 检查单元格是否有非黑色的字体颜色
        if customer_cell.font and customer_cell.font.color and customer_cell.font.color.rgb:
            if customer_cell.font.color.rgb != '00000000':
                source_font_color = customer_cell.font.color
        font_colors.append(source_font_color)
    df['font_color_obj'] = font_colors
    
    # 将'开票日期'列转换为日期时间格式，以便进行正确的日期排序
    # 如果有无法转换的日期，会将其设为NaT（Not a Time）
    df['开票日期'] = pd.to_datetime(df['开票日期'], format='%Y%m%d', errors='coerce')
    
    # 按“开票日期”列对整个数据集进行升序排序
    df.sort_values(by='开票日期', inplace=True)
    # 排序后重置索引，以保证后续遍历的顺序是正确的
    df.reset_index(drop=True, inplace=True)

    # 查找最晚的开票日期，以确定销售月份
    latest_date = df['开票日期'].max()
    sales_month = latest_date.strftime('%Y%m') if pd.notna(latest_date) else "未知年月"
    
    # 提取并清理规格名称，用于输出文件的头部
    spec_columns_raw = [col for col in df.columns if 'ml' in str(col)]
    spec_columns_clean = [col.replace('数量', '').replace('套', '').strip() for col in spec_columns_raw]
    spec_list_str = "、".join(spec_columns_clean)

    # --- 步骤 3: 创建输出工作簿和样式 ---
    wb_out = openpyxl.Workbook()
    sheet_out = wb_out.active
    
    # 写入固定的头部信息并设置格式
    sheet_out.merge_cells('A1:J1'); sheet_out['A1'] = '广州市云端白雁生物科技有限公司'; sheet_out['A1'].font = Font(name='宋体', size=20, bold=True); sheet_out['A1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet_out.merge_cells('A2:J2'); sheet_out['A2'] = f'销售记录表（{sales_month}）'; sheet_out['A2'].font = Font(name='宋体', size=16, bold=True); sheet_out['A2'].alignment = Alignment(horizontal='center', vertical='center')
    sheet_out.merge_cells('A3:J3'); sheet_out['A3'] = f'产品名称：清诺盐水鼻腔喷雾器                                   规格型号：{spec_list_str}'; sheet_out['A3'].font = Font(name='宋体', size=11); sheet_out['A3'].alignment = Alignment(horizontal='left', vertical='center')
    
    # 写入表头
    headers = ['序号', '购货单位名称', '购货单位地址', '联系方式', '数量(瓶)', '规格', '批号', '有效期至', '销售日期', '发货人']
    sheet_out.append(headers)
    
    # 定义将要使用的边框样式
    thin_border_side = Side(style='thin')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # 为表头行（第4行）的每个单元格设置格式
    for cell in sheet_out[4]: 
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # --- 步骤 4: 核心数据转换与写入 ---
    current_row_num = 5  # 输出表格的数据起始行号
    seq_counter = 1      # 序号计数器

    # 遍历已按开票日期排序的每一行源数据
    for index, source_row in df.iterrows():
        # a. 提取静态数据（地址、联系方式等）
        try:
            address = source_row.iloc[16] if pd.notna(source_row.iloc[16]) else ''
            contact = source_row.iloc[17] if pd.notna(source_row.iloc[17]) else ''
        except IndexError:
            address, contact = '', ''

        customer_name = source_row.get('客户名称', '')
        # 将日期对象格式化回YYYYMMDD字符串，以备写入
        sales_date = source_row['开票日期'].strftime('%Y%m%d') if pd.notna(source_row['开票日期']) else ""

        # b. 从数据行中直接获取已绑定的字体颜色信息
        source_font_color = source_row['font_color_obj']
        
        # c. 提取动态数据（批号、规格与数量）
        batch_raw = str(source_row.get('批号', ''))
        batches = [b.strip() for b in batch_raw.split('/') if b.strip()] if batch_raw else []
        
        specs_with_qty = []
        for i, spec_col in enumerate(spec_columns_raw):
            qty = source_row.get(spec_col)
            if pd.notna(qty) and qty != 0:
                specs_with_qty.append({'qty': qty, 'spec': spec_columns_clean[i]})

        # d. 根据批号和规格数量，应用拆分逻辑生成子行
        sub_rows = []
        num_batches, num_specs = len(batches), len(specs_with_qty)
        
        if num_specs == 0:
            continue

        if num_batches > 1 and num_specs == 1:
            spec_info = specs_with_qty[0]
            sub_rows.append({'qty': spec_info['qty'], 'spec': spec_info['spec'], 'batch': batches[0]})
            for i in range(1, num_batches):
                sub_rows.append({'qty': 0, 'spec': spec_info['spec'], 'batch': batches[i]})
        else:
            temp_batches = list(batches)
            for spec_info in specs_with_qty:
                batch_no = ""
                if num_batches == num_specs and temp_batches:
                    batch_no = temp_batches.pop(0)
                elif num_batches == 1:
                    batch_no = batches[0]
                elif num_batches > 0:
                    batch_no = batch_raw
                sub_rows.append({'qty': spec_info['qty'], 'spec': spec_info['spec'], 'batch': batch_no})

        if num_batches > 1 and num_specs > 1 and num_batches != num_specs:
            print(f"警告：源表第 {index+2} 行数据批号({num_batches})与规格({num_specs})不匹配。客户: {customer_name}")

        # e. 将生成的子行写入Excel并立即进行格式化
        group_start_row = current_row_num
        for sub_row in sub_rows:
            row_to_write = [
                '', customer_name, address, contact,
                sub_row['qty'], sub_row['spec'], sub_row['batch'],
                calculate_expiry_date(sub_row['batch']), sales_date, ''
            ]
            sheet_out.append(row_to_write)
            
            # 如果当前行有关联的特殊字体颜色，则应用它
            if source_font_color:
                sheet_out.cell(row=current_row_num, column=2).font = Font(color=source_font_color)
            
            # 为该行的所有单元格应用边框
            for col_idx in range(1, 11):
                sheet_out.cell(row=current_row_num, column=col_idx).border = thin_border
            
            current_row_num += 1

        # f. 对刚写入的区块进行单元格合并和对齐
        group_size = len(sub_rows)
        sheet_out.cell(row=group_start_row, column=1).value = seq_counter
        
        if group_size > 1:
            for col_idx in [1, 2, 3, 4, 9, 10]:
                sheet_out.merge_cells(start_row=group_start_row, start_column=col_idx, end_row=current_row_num - 1, end_column=col_idx)

        for col_idx in [1, 2, 3, 4, 9, 10]:
            cell = sheet_out.cell(row=group_start_row, column=col_idx)
            cell.alignment = Alignment(vertical='top', horizontal='center' if col_idx==1 else 'left')

        seq_counter += 1

    # --- 步骤 5: 设置列宽并保存文件 ---
    sheet_out.column_dimensions['A'].width = 5; sheet_out.column_dimensions['B'].width = 30; sheet_out.column_dimensions['C'].width = 45; sheet_out.column_dimensions['D'].width = 15; sheet_out.column_dimensions['E'].width = 10; sheet_out.column_dimensions['F'].width = 10; sheet_out.column_dimensions['G'].width = 10; sheet_out.column_dimensions['H'].width = 12; sheet_out.column_dimensions['I'].width = 12; sheet_out.column_dimensions['J'].width = 10

    try:
        wb_out.save(output_path)
        print(f"处理完成！文件已保存至: {output_path}")
    except PermissionError:
        print(f"错误：无法保存文件。请确保 {OUTPUT_FILENAME} 文件未被其他程序打开。")

# 当脚本作为主程序运行时，执行主处理函数
if __name__ == "__main__":
    process_sales_data()