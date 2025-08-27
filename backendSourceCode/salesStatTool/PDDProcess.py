import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import numpy as np

# --- 配置区 ---

# Pandas 显示选项
pd.set_option('future.no_silent_downcasting', True)

# 拼多多原始列名常量定义
PDD_COL_PRODUCT_NAME = '商品'
PDD_COL_ORDER_ID = '订单号'
PDD_COL_ORDER_STATUS = '订单状态'
PDD_COL_PRODUCT_TOTAL_PRICE = '商品总价(元)'
PDD_COL_STORE_DISCOUNT = '店铺优惠折扣(元)'
PDD_COL_PLATFORM_DISCOUNT = '平台优惠折扣(元)'
PDD_COL_USER_ACTUAL_PAYMENT = '用户实付金额(元)'
PDD_COL_MERCHANT_ACTUAL_RECEIPT = '商家实收金额(元)'
PDD_COL_QUANTITY = '商品数量(件)'
PDD_COL_SHIPPING_TIME = '发货时间'
PDD_COL_CONFIRM_RECEIPT_TIME = '确认收货时间'
PDD_COL_PRODUCT_ID = '商品id'
PDD_COL_STYLE_ID = '样式ID'
PDD_COL_PRODUCT_SPEC = '商品规格'
PDD_COL_AFTER_SALES_STATUS = '售后状态'
PDD_COL_LOGISTICS_NO = '快递单号'
PDD_COL_LOGISTICS_COMPANY = '快递公司'
PDD_COL_ORDER_TRANSACTION_TIME = '订单成交时间'

# 订单状态常量
STATUS_REFUND_SUCCESS = '退款成功'

# 输出到详情页的列定义，包含了商品ID和样式ID
DETAIL_SHEET_COLUMNS_PDD = [
    '订单号', '订单状态', '售后状态', '商品ID', '样式ID', '商品名称', '商品规格', '商品数量(件)',
    '用户实付金额(元)', '商家实收金额(元)', '订单成交时间', '发货时间', '确认收货时间',
    '快递单号', '快递公司'
]

# --- 内部功能函数 ---

def _prepare_and_validate_data(df):
    """
    验证输入DataFrame的结构，转换数值列，并进行初步的数据筛选和清洗。
    核心逻辑基于'样式ID'。
    """
    # 定义必须存在的数值列，用于后续计算
    numeric_cols = [
        PDD_COL_QUANTITY, PDD_COL_USER_ACTUAL_PAYMENT, PDD_COL_MERCHANT_ACTUAL_RECEIPT,
        PDD_COL_PRODUCT_TOTAL_PRICE, PDD_COL_STORE_DISCOUNT, PDD_COL_PLATFORM_DISCOUNT,
    ]
    # 遍历数值列，进行类型转换，无法转换的填充为0
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    # 过滤掉所有包含“取消”状态的订单
    df_no_cancel = df[~df[PDD_COL_ORDER_STATUS].str.contains('取消', na=False)].copy()
    
    # 将'样式ID'列的空值填充为“未知样式”，确保所有行都能被分组
    df_no_cancel[PDD_COL_STYLE_ID] = df_no_cancel[PDD_COL_STYLE_ID].fillna("未知样式")
    
    # 筛选出包含有效'样式ID'的行作为最终处理数据
    df_processed = df_no_cancel[df_no_cancel[PDD_COL_STYLE_ID].notna()].copy()
    
    if df_processed.empty:
        print("数据中没有找到未取消且包含有效样式ID的行。无法生成报告。")
        return None
    
    # 确保'样式ID'和'商品id'为字符串类型，防止科学计数法等问题
    df_processed[PDD_COL_STYLE_ID] = df_processed[PDD_COL_STYLE_ID].astype(str)
    df_processed[PDD_COL_PRODUCT_ID] = df_processed[PDD_COL_PRODUCT_ID].astype(str)

    return df_processed

def _format_df_for_detail(df_source, is_refund=False):
    """
    根据详情页列定义，格式化DataFrame，并根据是否为退款计算金额。
    此函数逐行提取原始信息，确保详情页中每条记录的准确性。
    """
    # 如果源数据为空，直接返回一个带表头的空DataFrame
    if df_source.empty: return pd.DataFrame(columns=DETAIL_SHEET_COLUMNS_PDD)
    
    # 创建一个新的DataFrame用于存放格式化后的数据
    df_target = pd.DataFrame()
    df_target['订单号'] = df_source.get(PDD_COL_ORDER_ID)
    df_target['订单状态'] = df_source.get(PDD_COL_ORDER_STATUS)
    df_target['售后状态'] = df_source.get(PDD_COL_AFTER_SALES_STATUS)
    df_target['商品ID'] = df_source.get(PDD_COL_PRODUCT_ID)
    df_target['样式ID'] = df_source.get(PDD_COL_STYLE_ID)
    df_target['商品名称'] = df_source.get(PDD_COL_PRODUCT_NAME) # 直接从源数据(df_source)的列中获取，保证每行都是真实的商品名称
    df_target['商品规格'] = df_source.get(PDD_COL_PRODUCT_SPEC)
    df_target['商品数量(件)'] = df_source.get(PDD_COL_QUANTITY)
    df_target['订单成交时间'] = df_source.get(PDD_COL_ORDER_TRANSACTION_TIME)
    df_target['发货时间'] = df_source.get(PDD_COL_SHIPPING_TIME)
    df_target['确认收货时间'] = df_source.get(PDD_COL_CONFIRM_RECEIPT_TIME)
    df_target['快递单号'] = df_source.get(PDD_COL_LOGISTICS_NO)
    df_target['快递公司'] = df_source.get(PDD_COL_LOGISTICS_COMPANY)

    # 根据是否为退款订单，计算金额（退款金额为负数）
    if not is_refund:
        # 收入计算：用户实付金额直接取值，商家实收=商品总价-店铺优惠
        df_target['用户实付金额(元)'] = df_source.get(PDD_COL_USER_ACTUAL_PAYMENT, 0)
        df_target['商家实收金额(元)'] = df_source.get(PDD_COL_PRODUCT_TOTAL_PRICE, 0) - df_source.get(PDD_COL_STORE_DISCOUNT, 0)
    else:
        # 支出计算：用户实付金额取负值，商家实收=-(用户实付+平台优惠)，不含平台优惠部分
        df_target['用户实付金额(元)'] = -df_source.get(PDD_COL_USER_ACTUAL_PAYMENT, 0)
        df_target['商家实收金额(元)'] = -(df_source.get(PDD_COL_USER_ACTUAL_PAYMENT, 0) + df_source.get(PDD_COL_PLATFORM_DISCOUNT, 0))

    # 按照预定义的列顺序重新排列，并填充空值为''
    return df_target.reindex(columns=DETAIL_SHEET_COLUMNS_PDD).fillna('')

def _aggregate_product_data(df_processed):
    """
    按'样式ID'对数据进行聚合，并将每个样式的数据划分为收入和两类退款（未发货/已发货）。
    """
    product_data_map = {}
    # 使用'样式ID'进行分组
    for style_id, group_df in df_processed.groupby(PDD_COL_STYLE_ID):
        # 从分组数据中提取公共信息（商品名、规格、商品ID）用于总结页和Sheet标题
        product_name = group_df[PDD_COL_PRODUCT_NAME].iloc[0] if not group_df[PDD_COL_PRODUCT_NAME].empty else "未知商品"
        product_spec = group_df[PDD_COL_PRODUCT_SPEC].iloc[0] if not group_df[PDD_COL_PRODUCT_SPEC].empty else "未知规格"
        
        # 筛选出所有退款成功的订单
        all_refunds_df = group_df[group_df[PDD_COL_AFTER_SALES_STATUS].str.contains(STATUS_REFUND_SUCCESS, na=False)]
        # 进一步细分未发货的退款
        unshipped_refund_df = all_refunds_df[all_refunds_df[PDD_COL_ORDER_STATUS].str.contains('未发货', na=False)]
        # 细分已发货的退款
        shipped_refund_df = all_refunds_df[all_refunds_df[PDD_COL_ORDER_STATUS].str.contains('已发货', na=False)]

        # 将每个样式的数据存入字典，键为样式ID
        product_data_map[style_id] = {
            'name': product_name, # 代表性名称
            'spec': product_spec, # 代表性规格
            'income_df_source': group_df, # 传递原始数据给详情页格式化函数
            'unshipped_refund_df_source': unshipped_refund_df,
            'shipped_refund_df_source': shipped_refund_df
        }
    return product_data_map

def _create_summary_sheet(wb, product_data_map):
    """
    在Excel工作簿中创建并填充销售总结页。
    """
    ws = wb.active
    ws.title = "销售总结"
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    current_row = 1

    # 内部函数，用于写入一个汇总区域（如收入、未发货退款等）
    def write_summary_section(title, df_source_key, headers, is_refund=False):
        nonlocal current_row
        ws.cell(row=current_row, column=1, value=title).font = bold_font
        current_row += 1
        ws.append(headers)
        for cell in ws[current_row]: cell.font = bold_font; cell.alignment = center_align
        current_row += 1
        
        total_qty, total_user_pay, total_receipt = 0, 0.0, 0.0
        sorted_ids = sorted(product_data_map.keys(), key=lambda x: (x == "未知样式", x))
        for s_id in sorted_ids:
            item = product_data_map[s_id]
            df = item[df_source_key]
            if not df.empty:
                # 为了计算总计，需要模拟格式化函数中的金额计算
                qty = df[PDD_COL_QUANTITY].sum()
                user_pay = df[PDD_COL_USER_ACTUAL_PAYMENT].sum()
                if is_refund:
                    receipt = -(df[PDD_COL_USER_ACTUAL_PAYMENT].sum() + df[PDD_COL_PLATFORM_DISCOUNT].sum())
                    user_pay = -user_pay
                else:
                    receipt = df[PDD_COL_PRODUCT_TOTAL_PRICE].sum() - df[PDD_COL_STORE_DISCOUNT].sum()

                ws.append([s_id, item['spec'], item['name'], qty, user_pay, receipt])
                total_qty += qty; total_user_pay += user_pay; total_receipt += receipt
                current_row += 1
        
        total_row_title = title.replace("各商品", "").replace("汇总", "总计").strip()
        ws.append([total_row_title, "", "", total_qty, total_user_pay, total_receipt])
        for cell in ws[current_row]: cell.font = bold_font
        current_row += 2
        return total_qty, total_user_pay, total_receipt

    income_qty, income_user, income_receipt = write_summary_section(
        "各商品收入汇总 (所有未取消订单)", 'income_df_source',
        ["样式ID", "商品规格", "商品名称", "总销售数量", "用户实付总额(参考)", "总销售额"]
    )
    unshipped_qty, unshipped_user, unshipped_receipt = write_summary_section(
        "各商品支出汇总 (未发货退款)", 'unshipped_refund_df_source',
        ["样式ID", "商品规格", "商品名称", "退款数量", "用户实付总额(退款)", "总退款额"], is_refund=True
    )
    shipped_qty, shipped_user, shipped_receipt = write_summary_section(
        "各商品支出汇总 (已发货退款)", 'shipped_refund_df_source',
        ["样式ID", "商品规格", "商品名称", "退款数量", "用户实付(退款)", "退款额"], is_refund=True
    )

    # 计算并写入两种口径的净总计
    net_qty1 = income_qty - unshipped_qty
    net_user1 = income_user + unshipped_user
    net_receipt1 = income_receipt + unshipped_receipt
    ws.cell(row=current_row, column=1, value="净总计(已发货退款订单按售出计算)").font = bold_font
    ws.cell(row=current_row, column=4, value=net_qty1).font = bold_font
    ws.cell(row=current_row, column=5, value=net_user1).font = bold_font
    ws.cell(row=current_row, column=6, value=net_receipt1).font = bold_font
    current_row += 1

    net_qty2 = income_qty - unshipped_qty - shipped_qty
    net_user2 = income_user + unshipped_user + shipped_user
    net_receipt2 = income_receipt + unshipped_receipt + shipped_receipt
    ws.cell(row=current_row, column=1, value="净总计(已发货退款订单按退款计算)").font = bold_font
    ws.cell(row=current_row, column=4, value=net_qty2).font = bold_font
    ws.cell(row=current_row, column=5, value=net_user2).font = bold_font
    ws.cell(row=current_row, column=6, value=net_receipt2).font = bold_font
    
    # 设置列宽和数字格式
    for col_letter, width in [('A', 35), ('B', 25), ('C', 60), ('D', 18), ('E', 22), ('F', 22)]:
        ws.column_dimensions[col_letter].width = width
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.column == 4: cell.number_format = '#,##0'
                if cell.column in [5, 6]: cell.number_format = '#,##0.00'

def _create_detail_sheets(wb, product_data_map):
    """
    为每个样式ID（SKU）创建并填充一个详情页。
    """
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')

    # 内部函数，用于在详情页中写入一个数据区域（如收入明细）
    def write_section(ws, df_source, title, total_title, is_refund=False):
        if df_source.empty: return
        # 格式化数据，确保每行信息准确
        df_formatted = _format_df_for_detail(df_source, is_refund=is_refund)
        
        ws.append([title]); ws[ws.max_row][0].font = bold_font
        ws.append(DETAIL_SHEET_COLUMNS_PDD)
        for cell in ws[ws.max_row]: cell.font = bold_font; cell.alignment = center_align
        for _, row_data in df_formatted.iterrows(): ws.append(row_data.tolist())
        
        # 计算并写入该区域的总计行
        qty_sum = df_formatted['商品数量(件)'].sum()
        user_pay_sum = df_formatted['用户实付金额(元)'].sum()
        receipt_sum = df_formatted['商家实收金额(元)'].sum()
        total_row_data = [total_title, "", "", "", "", "", "", qty_sum, user_pay_sum, receipt_sum]
        ws.append(total_row_data)
        for cell in ws[ws.max_row]: cell.font = bold_font
        ws.append([])

    # 遍历所有样式ID，创建对应的详情页
    sorted_ids = sorted(product_data_map.keys(), key=lambda x: (x == "未知样式", x))
    for s_id in sorted_ids:
        item = product_data_map[s_id]
        # 根据“样式ID_商品规格_商品标题”生成Sheet页名称，并做截断处理
        sheet_name_raw = f"{s_id}_{item['spec']}_{item['name']}"
        sheet_name = re.sub(r'[\\/\*\[\]\:?]', '_', sheet_name_raw)[:31]
        try: 
            ws = wb.create_sheet(sheet_name)
        except: 
            # 如果有命名冲突或其他错误，使用备用名称
            ws = wb.create_sheet(f"{s_id}_detail")
            
        # 写入收入和两类退款的明细数据
        write_section(ws, item['income_df_source'], "收入明细 (所有未取消订单)", "收入总计")
        write_section(ws, item['unshipped_refund_df_source'], "支出明细 (未发货退款)", "未发货退款总计", is_refund=True)
        write_section(ws, item['shipped_refund_df_source'], "支出明细 (已发货退款)", "已发货退款总计", is_refund=True)

        # 设置详情页的列宽
        for col_idx, width in enumerate([25, 20, 15, 25, 25, 60, 22, 15, 18, 18, 20, 20, 20, 25, 15], 1):
             ws.column_dimensions[get_column_letter(col_idx)].width = width

# --- 主处理函数 ---

def process_pdd_data(df_raw):
    """
    处理拼多多订单DataFrame，生成包含总结和明细的Excel Workbook对象。
    """
    if df_raw is None or df_raw.empty:
        print("拼多多处理错误：输入的DataFrame为空。")
        return None

    # 步骤1：数据准备与验证
    df_processed = _prepare_and_validate_data(df_raw)
    if df_processed is None: return None
    
    # 步骤2：按样式ID聚合数据
    product_data_map = _aggregate_product_data(df_processed)
    
    # 步骤3：创建Excel工作簿并生成页面
    wb = Workbook()
    _create_summary_sheet(wb, product_data_map)
    _create_detail_sheets(wb, product_data_map)
    
    # 确保“销售总结”页在第一个位置
    if "销售总结" in wb.sheetnames and wb.sheetnames[0] != "销售总结":
         wb.move_sheet("销售总结", -len(wb.sheetnames)+1)
        
    return wb

# ---- 主程序入口 (用于独立测试) ----
if __name__ == "__main__":
    TEST_INPUT_DIR = r"C:\Users\LENOVO\Desktop"
    TEST_OUTPUT_DIR = r"C:\Users\LENOVO\Desktop"
    TEST_FILENAME = "orders_export.csv"

    input_file = os.path.join(TEST_INPUT_DIR, TEST_FILENAME)

    if not os.path.exists(input_file):
        print(f"错误: 测试输入文件未找到于 '{os.path.abspath(input_file)}'")
    else:
        print(f"--- 独立测试: 读取文件 {TEST_FILENAME} ---")
        try:
            # 读取CSV文件，强制所有列为字符串类型
            df_test_raw = pd.read_csv(input_file, dtype=str, encoding='utf-8-sig')
            # 清理列名中的空格
            df_test_raw.columns = df_test_raw.columns.str.strip()
            # 统一处理空值和特定字符串为NaN
            for col in df_test_raw.columns:
                if df_test_raw[col].dtype == 'object':
                    df_test_raw[col] = df_test_raw[col].astype(str).str.strip().replace(
                        ['-', '--', '', 'None', 'nan', '#NULL!', None, 'null', '\t'], np.nan, regex=False)
        except Exception as e:
            print(f"测试中读取文件失败: {e}")
            df_test_raw = None

        if df_test_raw is not None:
            print("--- 独立测试: 调用 process_pdd_data ---")
            workbook_result = process_pdd_data(df_test_raw)
            
            if workbook_result:
                os.makedirs(TEST_OUTPUT_DIR, exist_ok=True)
                file_name_without_ext = os.path.splitext(TEST_FILENAME)[0]
                output_filename = f"PDD_output_{file_name_without_ext}.xlsx"
                output_file_path = os.path.join(TEST_OUTPUT_DIR, output_filename)
                
                try:
                    workbook_result.save(output_file_path)
                    print(f"\n脚本独立测试成功。输出文件位于: {output_file_path}")
                except Exception as e:
                    print(f"\n测试中保存文件失败: {e}")
            else:
                print("\n脚本独立测试失败，未能生成Workbook对象。")