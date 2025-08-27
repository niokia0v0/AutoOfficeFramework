import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import numpy as np

# --- 配置区 ---

# Pandas 显示选项
pd.set_option('future.no_silent_downcasting', True)

# 京东原始列名常量定义
JD_COL_ORDER_ID = '订单编号'
JD_COL_ORDER_STATUS = '订单状态'
JD_COL_PRODUCT_ID = '商品编号'
JD_COL_PRODUCT_NAME = '商品名称'
JD_COL_QUANTITY = '商品数量'
JD_COL_AMOUNT_DUE = '应结金额'
JD_COL_FEE_NAME = '费用名称'
JD_COL_DIRECTION = '收支方向'
JD_COL_AFTER_SALES_ID = '售后服务单号'

# 费用名称常量 (仅保留货款，其他费用不再统计)
FEE_NAME_GOODS = '货款'

# 订单状态常量
STATUS_COMPLETED = '已完成'

# 收支方向常量
DIRECTION_INCOME = '收入'
DIRECTION_EXPENSE = '支出'

# 输出到详情页的列定义 (已简化，移除不统计的费用的相关列)
DETAIL_SHEET_COLUMNS_JD = [
    '订单编号', '父单号', '订单状态', '订单下单时间', '订单完成时间', '售后服务单号', '售后退款时间',
    '商品编号', '商品名称', '商品数量', '费用名称', '应结金额',
    '收支方向', '结算状态', '预计结算时间', '账单生成时间', '到账时间', '商户订单号'
]

# --- 内部功能函数 ---

def _convert_numeric_columns(df):
    """
    将DataFrame中的指定列转换为数值类型，便于计算。
    """
    # 转换应结金额和商品数量为数值，无法转换的填充为0
    df[JD_COL_AMOUNT_DUE] = pd.to_numeric(df[JD_COL_AMOUNT_DUE], errors='coerce').fillna(0)
    df[JD_COL_QUANTITY] = pd.to_numeric(df[JD_COL_QUANTITY], errors='coerce').fillna(0)
    return df

def _filter_and_prepare_data(df_numeric):
    """
    筛选有效数据行，并进行必要的清洗和准备。
    - 筛选“已完成”状态的订单。
    - 清洗作为分组键的“商品名称”列。
    """
    # 仅处理状态为“已完成”的订单
    df_completed = df_numeric[df_numeric[JD_COL_ORDER_STATUS] == STATUS_COMPLETED].copy()
    
    # 确保售后相关列存在，以便后续筛选退款单
    for col in [JD_COL_AFTER_SALES_ID, '售后退款时间']:
        if col not in df_completed.columns:
            df_completed[col] = np.nan
            
    # 清洗商品名称列：去除首尾空格，并将空值替换为指定字符串
    df_completed[JD_COL_PRODUCT_NAME] = df_completed[JD_COL_PRODUCT_NAME].str.strip()
    df_completed[JD_COL_PRODUCT_NAME] = df_completed[JD_COL_PRODUCT_NAME].fillna("未知商品标题")

    # 筛选出包含有效商品名称的行
    df_processed = df_completed[df_completed[JD_COL_PRODUCT_NAME].notna()].copy()
    
    if df_processed.empty:
        print(f"数据中没有找到“已完成”且包含有效商品名称('{JD_COL_PRODUCT_NAME}')的行。无法生成报告。")
        return None
        
    return df_processed

def _aggregate_product_data(df_processed):
    """
    按'商品名称'聚合数据，仅计算货款的收入与支出。
    """
    product_summary = {}
    # 使用'商品名称'作为分组键，实现按规格（标题）聚合
    for product_name, group in df_processed.groupby(JD_COL_PRODUCT_NAME):
        # 从分组中提取该规格对应的商品编号（通常取第一个）
        product_id = group[JD_COL_PRODUCT_ID].dropna().iloc[0] if not group[JD_COL_PRODUCT_ID].dropna().empty else "未知编号"

        # 筛选货款收入的行
        sales_group = group[(group[JD_COL_FEE_NAME] == FEE_NAME_GOODS) & (group[JD_COL_DIRECTION] == DIRECTION_INCOME)]
        
        # 筛选货款支出的行（即退款），并要求有售后服务单号
        returns_group = group[
            (group[JD_COL_FEE_NAME] == FEE_NAME_GOODS) & 
            (group[JD_COL_DIRECTION] == DIRECTION_EXPENSE) & 
            (group[JD_COL_AFTER_SALES_ID].notna()) & 
            (group[JD_COL_AFTER_SALES_ID] != '')
        ]
        
        # 将聚合结果存入字典，键为商品名称
        product_summary[str(product_name)] = {
            'prod_id': product_id,
            'sales_quantity': sales_group[JD_COL_QUANTITY].sum(),
            'sales_amount': sales_group[JD_COL_AMOUNT_DUE].sum(),
            'return_quantity': returns_group[JD_COL_QUANTITY].sum(),
            'return_amount': returns_group[JD_COL_AMOUNT_DUE].sum(), # 金额为负数
            'sales_detail_df': sales_group,
            'returns_detail_df': returns_group,
        }
    return product_summary

def _create_summary_sheet(wb, product_summary):
    """
    在工作簿中创建并填充销售总结页，采用销售、退款、总计三段式布局。
    """
    ws = wb.active
    ws.title = "销售总结"
    bold_font = Font(bold=True)
    
    # --- 1. 销售汇总区域 ---
    headers = ["商品编号", "商品名称", "销售数量", "销售额"]
    ws.append(headers)
    for cell in ws[1]: cell.font = bold_font
    
    sorted_names = sorted(product_summary.keys())
    for name in sorted_names:
        item = product_summary[name]
        # 只添加有销售额的行
        if item['sales_amount'] != 0:
            ws.append([item['prod_id'], name, item['sales_quantity'], item['sales_amount']])

    sales_end_row = ws.max_row
    current_row = sales_end_row + 1
    # 添加销售总计行
    ws.cell(row=current_row, column=1, value="总计 (不含退款)").font = bold_font
    ws.cell(row=current_row, column=3, value=f"=SUM(C2:C{sales_end_row})").font = bold_font
    ws.cell(row=current_row, column=4, value=f"=SUM(D2:D{sales_end_row})").font = bold_font
    sales_total_row = current_row

    # --- 2. 退款明细区域 ---
    current_row += 2
    ws.cell(row=current_row, column=1, value="退款商品明细").font = bold_font
    current_row += 1
    
    refund_start_row = current_row
    has_refunds = False
    for name in sorted_names:
        item = product_summary[name]
        if item['return_quantity'] > 0 or item['return_amount'] != 0:
            has_refunds = True
            ws.append([item['prod_id'], name, item['return_quantity'], item['return_amount']])
    
    refund_end_row = ws.max_row
    current_row = refund_end_row + 1
    # 添加退款总计行
    ws.cell(row=current_row, column=1, value="总计退款").font = bold_font
    if has_refunds:
        ws.cell(row=current_row, column=3, value=f"=SUM(C{refund_start_row}:C{refund_end_row})").font = bold_font
        ws.cell(row=current_row, column=4, value=f"=SUM(D{refund_start_row}:D{refund_end_row})").font = bold_font
    else:
        # 如果没有退款，则填0
        ws.cell(row=current_row, column=3, value=0).font = bold_font
        ws.cell(row=current_row, column=4, value=0).font = bold_font
    refund_total_row = current_row

    # --- 3. 最终总计 ---
    current_row += 2
    ws.cell(row=current_row, column=1, value="总计 (计算退款)").font = bold_font
    # 净数量 = 销售数量 - 退款数量 (假设退款数量是正数)
    # 净金额 = 销售额 + 退款额 (因为退款金额本身是负数)
    ws.cell(row=current_row, column=3, value=f"=C{sales_total_row}-C{refund_total_row}").font = bold_font
    ws.cell(row=current_row, column=4, value=f"=D{sales_total_row}+D{refund_total_row}").font = bold_font
    
    # --- 格式化 ---
    ws.column_dimensions['A'].width = 25; ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 18; ws.column_dimensions['D'].width = 18
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                if cell.column == 3: cell.number_format = '#,##0'
                else: cell.number_format = '#,##0.00'

def _create_detail_sheets(wb, product_summary):
    """
    为每个商品名称（规格）创建并填充详情页。
    """
    for name, item in product_summary.items():
        if item['sales_detail_df'].empty and item['returns_detail_df'].empty: continue

        # --- Sheet页命名逻辑 ---
        # 1. 拼接原始长名称
        sheet_name_raw = f"{item['prod_id']}_{name}"
        # 2. 立即清理所有Excel不支持的特殊字符
        base_name = re.sub(r'[\\/\*\[\]\:?]', '_', sheet_name_raw)
        
        # 3. 对清理后的名称进行长度检查和截断
        if len(base_name) > 31:
            id_prefix = f"{item['prod_id']}_..."
            # 重新清理一次商品名本身，以确保截断源是干净的
            clean_name = re.sub(r'[\\/\*\[\]\:?]', '_', name)
            available_len = 31 - len(id_prefix) - 4 # 预留空间给序号
            truncated_name = clean_name[-available_len:] if available_len > 0 else ""
            base_name = f"{id_prefix}{truncated_name}"

        sheet_name = base_name
        counter = 1
        # 4. 防重名处理
        while sheet_name in wb.sheetnames:
            counter += 1
            suffix = f"({counter})"
            truncated_base = base_name[:31-len(suffix)]
            sheet_name = f"{truncated_base}{suffix}"

        try:
            ws = wb.create_sheet(sheet_name)
        except Exception as e:
            print(f"警告: 创建Sheet页 '{sheet_name}' 失败: {e}。将使用备用名称。")
            ws = wb.create_sheet(f"{item['prod_id']}_detail")
            
        header_written = False
        def write_df_section(df, title, qty, amt):
            nonlocal header_written
            if df.empty: return
            ws.append([title]); ws[ws.max_row][0].font = Font(bold=True)
            if not header_written:
                ws.append(DETAIL_SHEET_COLUMNS_JD); header_written = True
                for cell in ws[ws.max_row]: cell.font = Font(bold=True)
            
            for r in df.reindex(columns=DETAIL_SHEET_COLUMNS_JD).fillna('').itertuples(index=False):
                ws.append(list(r))
            
            qty_col_idx = DETAIL_SHEET_COLUMNS_JD.index(JD_COL_QUANTITY)
            amt_col_idx = DETAIL_SHEET_COLUMNS_JD.index(JD_COL_AMOUNT_DUE)
            total_row = ["总计"] + [""] * (len(DETAIL_SHEET_COLUMNS_JD) - 1)
            total_row[qty_col_idx] = qty
            total_row[amt_col_idx] = amt
            ws.append(total_row)
            
            for cell in ws[ws.max_row]: cell.font = Font(bold=True)
            ws.append([])
        
        write_df_section(item['sales_detail_df'], "销售明细 (货款)", item['sales_quantity'], item['sales_amount'])
        write_df_section(item['returns_detail_df'], "退款明细 (货款)", item['return_quantity'], item['return_amount'])

# --- 主处理函数 ---

def process_jingdong_data(df_raw):
    """
    处理京东结算DataFrame，生成包含总结和明细的Excel Workbook对象。
    """
    if df_raw is None or df_raw.empty:
        print("京东处理错误：输入的DataFrame为空。")
        return None
        
    df_numeric = _convert_numeric_columns(df_raw.copy())
    df_processed = _filter_and_prepare_data(df_numeric)
    if df_processed is None: return None
        
    product_summary = _aggregate_product_data(df_processed)
    
    wb = Workbook()
    _create_summary_sheet(wb, product_summary)
    _create_detail_sheets(wb, product_summary)
    
    if "销售总结" in wb.sheetnames and wb.sheetnames[0] != "销售总结":
        wb.move_sheet("销售总结", -len(wb.sheetnames)+1)
        
    return wb

# ---- 主程序入口 (用于独立测试) ----
if __name__ == "__main__":
    TEST_INPUT_DIR = r"C:\Users\LENOVO\Desktop"
    TEST_OUTPUT_DIR = r"C:\Users\LENOVO\Desktop"
    TEST_FILENAME = "订单结算明细对账.csv"

    input_file = os.path.join(TEST_INPUT_DIR, TEST_FILENAME)

    if not os.path.exists(input_file):
        print(f"错误: 测试输入文件未找到于 '{os.path.abspath(input_file)}'")
    else:
        print(f"--- 独立测试: 读取文件 {TEST_FILENAME} ---")
        try:
            df_test_raw = pd.read_csv(input_file, dtype=str, na_values=['--'], keep_default_na=True, encoding='utf-8-sig')
            df_test_raw.columns = [col.strip() for col in df_test_raw.columns]
            for col in df_test_raw.columns:
                if df_test_raw[col].dtype == 'object':
                    df_test_raw[col] = df_test_raw[col].str.strip().replace(['--', '', 'None', 'nan', np.nan], np.nan, regex=False)
        except Exception as e:
            print(f"测试中读取文件失败: {e}")
            df_test_raw = None

        if df_test_raw is not None:
            print("--- 独立测试: 调用 process_jingdong_data ---")
            workbook_result = process_jingdong_data(df_test_raw)
            
            if workbook_result:
                os.makedirs(TEST_OUTPUT_DIR, exist_ok=True)
                file_name_without_ext = os.path.splitext(TEST_FILENAME)[0]
                output_filename = f"JD_output_{file_name_without_ext}.xlsx"
                output_file_path = os.path.join(TEST_OUTPUT_DIR, output_filename)
                
                try:
                    workbook_result.save(output_file_path)
                    print(f"\n脚本独立测试成功。输出文件位于: {output_file_path}")
                except Exception as e:
                    print(f"\n测试中保存文件失败: {e}")
            else:
                print("\n脚本独立测试失败，未能生成Workbook对象。")