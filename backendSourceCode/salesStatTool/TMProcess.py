import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import numpy as np

# --- 配置区 ---
pd.set_option('future.no_silent_downcasting', True)
TMALL_COL_SUB_ORDER_ID = '子订单编号'
TMALL_COL_MAIN_ORDER_ID = '主订单编号'
TMALL_COL_PRODUCT_NAME = '商品标题'
TMALL_COL_UNIT_PRICE = '商品价格'
TMALL_COL_QUANTITY = '购买数量'
TMALL_COL_PRODUCT_ATTRIBUTES = '商品属性'
TMALL_COL_ORDER_STATUS = '订单状态'
TMALL_COL_ACTUAL_PAYMENT = '买家实付金额'
TMALL_COL_REFUND_STATUS = '退款状态'
TMALL_COL_REFUND_AMOUNT = '退款金额'
TMALL_COL_ORDER_CREATE_TIME = '订单创建时间'
TMALL_COL_ORDER_PAY_TIME = '订单付款时间'
TMALL_COL_PRODUCT_ID = '商品ID'
TMALL_COL_SHIPPING_TIME = '发货时间'
TMALL_COL_LOGISTICS_NO = '物流单号'
TMALL_COL_LOGISTICS_COMPANY = '物流公司'
STATUS_TRADE_SUCCESS = '交易成功'
DETAIL_SHEET_COLUMNS_TMALL = [
    '订单编号', '子订单编号', '订单状态', '退款状态', '商品编号', '商品名称',
    '商品属性', '商品价格', '商品数量', '应结金额', '订单创建时间', '订单付款时间',
    '发货时间', '物流单号', '物流公司'
]

# --- 内部功能函数 ---

def _prepare_and_validate_data(df):
    """验证数据，转换数值列，并进行最终的数据准备。"""
    critical_cols = [
        TMALL_COL_PRODUCT_ID, TMALL_COL_ORDER_STATUS, TMALL_COL_QUANTITY,
        TMALL_COL_ACTUAL_PAYMENT, TMALL_COL_REFUND_AMOUNT, TMALL_COL_PRODUCT_ATTRIBUTES
    ]
    for col in critical_cols:
        if col not in df.columns:
            print(f"错误: 核心逻辑所需列 '{col}' 在文件中未找到。脚本无法继续。")
            return None

    numeric_cols = {
        TMALL_COL_QUANTITY: 0, TMALL_COL_ACTUAL_PAYMENT: 0.0,
        TMALL_COL_REFUND_AMOUNT: 0.0, TMALL_COL_UNIT_PRICE: 0.0
    }
    for col, fill_value in numeric_cols.items():
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(fill_value)
        elif col == TMALL_COL_UNIT_PRICE:
            print(f"警告: 列 '{col}' 在文件中未找到，将创建并填充默认值0.0。")
            df[col] = 0.0

    df[TMALL_COL_PRODUCT_ID] = df[TMALL_COL_PRODUCT_ID].astype(str).replace('nan', np.nan)
    df_processed = df[df[TMALL_COL_PRODUCT_ID].notna()].copy()
    
    if df_processed.empty:
        print(f"数据中没有找到包含有效商品ID('{TMALL_COL_PRODUCT_ID}')的行。无法生成报告。")
        return None
        
    return df_processed

def _aggregate_product_data(df_processed):
    """按商品ID聚合数据，计算各商品的收入、支出和明细。"""
    product_data_map = {}
    
    successful_trades_df = df_processed[df_processed[TMALL_COL_ORDER_STATUS] == STATUS_TRADE_SUCCESS]
    successful_trades_total = successful_trades_df[TMALL_COL_ACTUAL_PAYMENT].sum()

    for product_id, group_df in df_processed.groupby(TMALL_COL_PRODUCT_ID):
        product_name = group_df[TMALL_COL_PRODUCT_NAME].dropna().iloc[0] if not group_df[TMALL_COL_PRODUCT_NAME].dropna().empty else "未知商品"
        income_df = group_df.copy()
        expenditure_df = group_df[group_df[TMALL_COL_ORDER_STATUS] != STATUS_TRADE_SUCCESS].copy()

        product_data_map[str(product_id)] = {
            'name': product_name,
            'income_qty': income_df[TMALL_COL_QUANTITY].sum(),
            'income_amount': income_df[TMALL_COL_ACTUAL_PAYMENT].sum(),
            'expenditure_qty': expenditure_df[TMALL_COL_QUANTITY].sum(),
            'expenditure_amount': -expenditure_df[TMALL_COL_REFUND_AMOUNT].sum(),
            'detail_income_df': income_df,
            'detail_expenditure_df': expenditure_df
        }
    return product_data_map, successful_trades_total

def _create_summary_sheet(wb, product_data_map, successful_trades_total):
    """在工作簿中创建并填充销售总结页，确保数字格式正确。"""
    ws = wb.active
    ws.title = "销售总结"
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    current_row = 1

    # --- 收入汇总 ---
    ws.cell(row=current_row, column=1, value="各商品收入汇总 (所有订单)").font = bold_font
    current_row += 1
    ws.append(["商品编号", "商品名称", "总销售数量", "总销售额(收入)"])
    for cell in ws[current_row]: cell.font = bold_font; cell.alignment = center_align
    current_row += 1
    
    grand_total_income_qty, grand_total_income_amt = 0, 0.0
    sorted_ids = sorted(product_data_map.keys())
    for prod_id in sorted_ids:
        item = product_data_map[prod_id]
        ws.append([prod_id, item['name'], item['income_qty'], item['income_amount']])
        ws.cell(row=current_row, column=3).number_format = '#,##0'
        ws.cell(row=current_row, column=4).number_format = '#,##0.00'
        grand_total_income_qty += item['income_qty']
        grand_total_income_amt += item['income_amount']
        current_row += 1
    
    ws.append(["总计收入", "", grand_total_income_qty, grand_total_income_amt])
    for cell in ws[current_row]: cell.font = bold_font
    ws.cell(row=current_row, column=3).number_format = '#,##0'
    ws.cell(row=current_row, column=4).number_format = '#,##0.00'
    current_row += 2

    # --- 支出汇总 ---
    ws.cell(row=current_row, column=1, value="各商品支出汇总 (非交易成功订单)").font = bold_font
    current_row += 1
    ws.append(["商品编号", "商品名称", "未成功订单商品数量", "总退款额(支出)"])
    for cell in ws[current_row]: cell.font = bold_font; cell.alignment = center_align
    current_row += 1

    grand_total_exp_qty, grand_total_exp_amt = 0, 0.0
    for prod_id in sorted_ids:
        item = product_data_map[prod_id]
        if item['expenditure_qty'] > 0 or item['expenditure_amount'] != 0:
            ws.append([prod_id, item['name'], item['expenditure_qty'], item['expenditure_amount']])
            ws.cell(row=current_row, column=3).number_format = '#,##0'
            ws.cell(row=current_row, column=4).number_format = '#,##0.00'
            grand_total_exp_qty += item['expenditure_qty']
            grand_total_exp_amt += item['expenditure_amount']
            current_row += 1

    ws.append(["总计支出", "", grand_total_exp_qty, grand_total_exp_amt])
    for cell in ws[current_row]: cell.font = bold_font
    ws.cell(row=current_row, column=3).number_format = '#,##0'
    ws.cell(row=current_row, column=4).number_format = '#,##0.00'
    current_row += 2

    # --- 净总计与最终总额 ---
    ws.cell(row=current_row, column=1, value="净总计").font = bold_font
    cell_net_qty = ws.cell(row=current_row, column=3, value=grand_total_income_qty - grand_total_exp_qty)
    cell_net_qty.font = bold_font; cell_net_qty.number_format = '#,##0'
    cell_net_amt = ws.cell(row=current_row, column=4, value=grand_total_income_amt + grand_total_exp_amt)
    cell_net_amt.font = bold_font; cell_net_amt.number_format = '#,##0.00'
    current_row += 1
    
    ws.cell(row=current_row, column=1, value="买家实付款总额(交易成功订单)").font = bold_font
    cell_success_amt = ws.cell(row=current_row, column=4, value=successful_trades_total)
    cell_success_amt.font = bold_font; cell_success_amt.number_format = '#,##0.00'
    
    # --- 列宽 ---
    for col_letter, width in [('A', 35), ('B', 60), ('C', 20), ('D', 20)]:
        ws.column_dimensions[col_letter].width = width

def _create_detail_sheets(wb, product_data_map):
    """为每个商品创建并填充详情页，确保数字格式正确。"""
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')

    def format_for_detail(df, amount_col, is_negative=False):
        if df.empty: return pd.DataFrame(columns=DETAIL_SHEET_COLUMNS_TMALL)
        detail = pd.DataFrame()
        detail['订单编号'] = df.get(TMALL_COL_MAIN_ORDER_ID)
        detail['子订单编号'] = df.get(TMALL_COL_SUB_ORDER_ID)
        detail['订单状态'] = df.get(TMALL_COL_ORDER_STATUS)
        detail['退款状态'] = df.get(TMALL_COL_REFUND_STATUS)
        detail['商品编号'] = df.get(TMALL_COL_PRODUCT_ID).astype(str)
        detail['商品名称'] = df.get(TMALL_COL_PRODUCT_NAME)
        detail['商品属性'] = df.get(TMALL_COL_PRODUCT_ATTRIBUTES)
        detail['商品价格'] = df.get(TMALL_COL_UNIT_PRICE)
        detail['商品数量'] = df.get(TMALL_COL_QUANTITY)
        amounts = df.get(amount_col, 0.0)
        detail['应结金额'] = -amounts if is_negative else amounts
        detail['订单创建时间'] = df.get(TMALL_COL_ORDER_CREATE_TIME)
        detail['订单付款时间'] = df.get(TMALL_COL_ORDER_PAY_TIME)
        detail['发货时间'] = df.get(TMALL_COL_SHIPPING_TIME)
        detail['物流单号'] = df.get(TMALL_COL_LOGISTICS_NO)
        detail['物流公司'] = df.get(TMALL_COL_LOGISTICS_COMPANY)
        return detail.reindex(columns=DETAIL_SHEET_COLUMNS_TMALL).fillna('')

    for prod_id in sorted(product_data_map.keys()):
        item = product_data_map[prod_id]
        
        sheet_name_raw = f"{prod_id}_{item['name']}"
        sheet_name = re.sub(r'[\\/\*\[\]\:?]', '_', sheet_name_raw)[:31]
        try: ws = wb.create_sheet(sheet_name)
        except: ws = wb.create_sheet(f"{prod_id}_detail")

        ws.append(DETAIL_SHEET_COLUMNS_TMALL)
        for cell in ws[1]: cell.font = bold_font; cell.alignment = center_align
        
        income_detail_df = format_for_detail(item['detail_income_df'], TMALL_COL_ACTUAL_PAYMENT)
        for r in income_detail_df.itertuples(index=False): ws.append(list(r))
        
        if not income_detail_df.empty:
            ws.append(["收入总计", "", "", "", "", "", "", "", item['income_qty'], item['income_amount']])
            row_idx = ws.max_row
            for cell in ws[row_idx]: cell.font = bold_font
            ws.cell(row=row_idx, column=9).number_format = '#,##0'
            ws.cell(row=row_idx, column=10).number_format = '#,##0.00'
            ws.append([])

        exp_detail_df = format_for_detail(item['detail_expenditure_df'], TMALL_COL_REFUND_AMOUNT, True)
        if not exp_detail_df.empty:
            for r in exp_detail_df.itertuples(index=False): ws.append(list(r))
            ws.append(["支出总计", "", "", "", "", "", "", "", item['expenditure_qty'], item['expenditure_amount']])
            row_idx = ws.max_row
            for cell in ws[row_idx]: cell.font = bold_font
            ws.cell(row=row_idx, column=9).number_format = '#,##0'
            ws.cell(row=row_idx, column=10).number_format = '#,##0.00'

        for i, col_title in enumerate(DETAIL_SHEET_COLUMNS_TMALL, 1):
            max_len = max((len(str(c.value)) for c in ws[get_column_letter(i)] if c.value is not None), default=0)
            adjusted_width = min(max(max_len + 5, len(col_title) + 5, 12), 60)
            if col_title == "商品名称": adjusted_width = 70
            elif col_title == "商品属性": adjusted_width = 40
            ws.column_dimensions[get_column_letter(i)].width = adjusted_width

# --- 主处理函数 (新架构) ---

def process_tmall_data(df_raw):
    """处理天猫订单DataFrame，生成包含总结和明细的Excel Workbook对象。"""
    if df_raw is None or df_raw.empty:
        print("天猫处理错误：输入的DataFrame为空。")
        return None
        
    df_processed = _prepare_and_validate_data(df_raw.copy())
    if df_processed is None: return None
        
    product_data_map, successful_trades_total = _aggregate_product_data(df_processed)
    
    wb = Workbook()
    _create_summary_sheet(wb, product_data_map, successful_trades_total)
    _create_detail_sheets(wb, product_data_map)
    
    if "销售总结" in wb.sheetnames and wb.sheetnames[0] != "销售总结":
        wb.move_sheet("销售总结", -len(wb.sheetnames) + 1)
        
    return wb

# ---- 主程序入口 (用于独立测试) ----
if __name__ == "__main__":
    TEST_INPUT_DIR = r"C:\Users\LENOVO\Desktop"
    TEST_OUTPUT_DIR = r"C:\Users\LENOVO\Desktop"
    TEST_FILENAME = "ExportOrderList1145141919810.xlsx"

    input_file = os.path.join(TEST_INPUT_DIR, TEST_FILENAME)

    if not os.path.exists(input_file):
        print(f"错误: 测试输入文件未找到于 '{os.path.abspath(input_file)}'")
    else:
        print(f"--- 独立测试: 读取文件 {TEST_FILENAME} ---")
        try:
            df_test_raw = pd.read_excel(input_file, dtype=str, engine='openpyxl')
            df_test_raw.columns = df_test_raw.columns.str.strip()
            for col in df_test_raw.columns:
                if df_test_raw[col].dtype == 'object':
                    df_test_raw[col] = df_test_raw[col].str.strip().replace(
                        ['--', '', 'None', 'nan', '#NULL!', None], np.nan, regex=False
                    )
        except Exception as e:
            print(f"测试中读取文件失败: {e}")
            df_test_raw = None

        if df_test_raw is not None:
            print("--- 独立测试: 调用 process_tmall_data ---")
            workbook_result = process_tmall_data(df_test_raw)
            
            if workbook_result:
                os.makedirs(TEST_OUTPUT_DIR, exist_ok=True)
                file_name_without_ext = os.path.splitext(TEST_FILENAME)[0]
                output_filename = f"TM_output_{file_name_without_ext}.xlsx"
                output_file_path = os.path.join(TEST_OUTPUT_DIR, output_filename)
                
                try:
                    workbook_result.save(output_file_path)
                    print(f"\n脚本独立测试成功。输出文件位于: {output_file_path}")
                except Exception as e:
                    print(f"\n测试中保存文件失败: {e}")
            else:
                print("\n脚本独立测试失败，未能生成Workbook对象。")