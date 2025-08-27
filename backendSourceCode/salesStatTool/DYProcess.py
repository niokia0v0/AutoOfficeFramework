import pandas as pd
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import numpy as np

# --- 配置区 ---

# Pandas 显示选项
pd.set_option('future.no_silent_downcasting', True)

# 抖店原始列名常量定义
DY_COL_MAIN_ORDER_ID = '主订单编号'
DY_COL_PRODUCT_NAME_DESC = '选购商品'
DY_COL_PRODUCT_ID = '商品ID'
DY_COL_QUANTITY = '商品数量'
DY_COL_UNIT_PRICE = '商品金额' # 这是单价
DY_COL_ORDER_SUBMIT_TIME = '订单提交时间'
DY_COL_PAY_COMPLETE_TIME = '支付完成时间'
DY_COL_ORDER_COMPLETE_TIME = '订单完成时间'
DY_COL_ORDER_STATUS = '订单状态'
DY_COL_CANCEL_REASON = '取消原因'
DY_COL_AFTER_SALES_STATUS = '售后状态'

# 订单状态常量
STATUS_COMPLETED = '已完成'

# 新增计算列名
CALC_COL_PAYABLE = '应付款'

# 输出到详情页的列定义
DETAIL_SHEET_COLUMNS_DY = [
    '订单编号', '订单状态', '售后状态', '取消原因', '商品编号', '商品名称',
    '商品单价', '商品数量', CALC_COL_PAYABLE, '订单提交时间', '支付完成时间', '订单完成时间'
]

# --- 内部功能函数 ---

def _prepare_and_validate_data(df):
    """
    验证数据，转换数值列，并进行最终的数据准备。
    此函数接收一个已经经过基础清洗的DataFrame。
    
    Args:
        df (pd.DataFrame): 原始数据DataFrame。
        
    Returns:
        pd.DataFrame: 准备好的DataFrame，如果关键列缺失或无有效数据则返回None。
    """
    # 验证关键列是否存在
    critical_cols = {
        DY_COL_PRODUCT_ID: "商品ID", DY_COL_PRODUCT_NAME_DESC: "选购商品",
        DY_COL_ORDER_STATUS: "订单状态", DY_COL_QUANTITY: "商品数量", 
        DY_COL_UNIT_PRICE: "商品金额"
    }
    for col_const, col_display in critical_cols.items():
        if col_const not in df.columns:
            print(f"错误: 核心逻辑所需列 '{col_display}' 在文件中未找到。脚本无法继续。")
            return None

    # 数值化转换
    numeric_cols = {DY_COL_QUANTITY: 0, DY_COL_UNIT_PRICE: 0.0}
    for col, fill_value in numeric_cols.items():
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(fill_value)

    # 创建应付款计算列 (单价 * 数量)
    df[CALC_COL_PAYABLE] = df[DY_COL_UNIT_PRICE] * df[DY_COL_QUANTITY]

    # 清洗作为分组键的“选购商品”列
    df[DY_COL_PRODUCT_NAME_DESC] = df[DY_COL_PRODUCT_NAME_DESC].str.strip()
    df[DY_COL_PRODUCT_NAME_DESC] = df[DY_COL_PRODUCT_NAME_DESC].fillna("未知商品标题")
    
    # 确保商品ID为字符串类型，以便后续提取
    df[DY_COL_PRODUCT_ID] = df[DY_COL_PRODUCT_ID].astype(str).replace('nan', '')

    # 筛选出包含有效商品标题的行
    df_processed = df[df[DY_COL_PRODUCT_NAME_DESC].notna()].copy()
    
    if df_processed.empty:
        print(f"数据中没有找到包含有效商品标题('{DY_COL_PRODUCT_NAME_DESC}')的行。无法生成报告。")
        return None
        
    return df_processed

def _aggregate_product_data(df_processed):
    """
    按'选购商品'（商品标题）聚合数据，计算各商品的收入、支出和明细。
    
    Args:
        df_processed (pd.DataFrame): 准备好的数据。
        
    Returns:
        dict: 一个包含所有商品聚合信息的字典 (product_data_map)。
    """
    product_data_map = {}
    # 使用'选购商品'列作为分组键
    for product_name, group_df in df_processed.groupby(DY_COL_PRODUCT_NAME_DESC):
        # 从分组数据中提取代表性的商品ID
        product_id = group_df[DY_COL_PRODUCT_ID].dropna().iloc[0] if not group_df[DY_COL_PRODUCT_ID].dropna().empty else "未知编号"

        # 收入数据为所有订单
        income_df = group_df.copy()
        # 支出数据为所有非“已完成”状态的订单
        expenditure_df = group_df[group_df[DY_COL_ORDER_STATUS] != STATUS_COMPLETED].copy()
        
        # 将聚合信息存入字典，键为商品名称
        product_data_map[product_name] = {
            'prod_id': product_id,
            'income_qty': income_df[DY_COL_QUANTITY].sum(),
            'income_amount': income_df[CALC_COL_PAYABLE].sum(),
            'expenditure_qty': expenditure_df[DY_COL_QUANTITY].sum(),
            'expenditure_amount': -(expenditure_df[CALC_COL_PAYABLE].sum()), # 支出金额为负
            'detail_income_df': income_df,
            'detail_expenditure_df': expenditure_df,
        }
    return product_data_map

def _create_summary_sheet(wb, product_data_map):
    """在工作簿中创建并填充销售总结页。"""
    ws = wb.active
    ws.title = "销售总结"
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    
    current_row = 1
    
    # --- 收入汇总 ---
    ws.cell(row=current_row, column=1, value="各商品收入汇总 (全部订单)").font = bold_font
    current_row += 1
    # 更新表头以同时展示商品编号和名称
    income_headers = ["商品编号", "商品名称", "总销售数量", "总应付金额"]
    ws.append(income_headers)
    for cell in ws[current_row]: cell.font = bold_font; cell.alignment = center_align
    current_row += 1

    grand_total_income_qty, grand_total_income_amt = 0, 0.0
    sorted_product_names = sorted(product_data_map.keys())
    
    for name in sorted_product_names:
        item = product_data_map[name]
        ws.append([item['prod_id'], name, item['income_qty'], item['income_amount']])
        grand_total_income_qty += item['income_qty']
        grand_total_income_amt += item['income_amount']
        current_row += 1
        
    ws.append(["总计收入", "", grand_total_income_qty, grand_total_income_amt])
    for cell in ws[current_row]: cell.font = bold_font
    current_row += 2
    
    # --- 支出汇总 ---
    ws.cell(row=current_row, column=1, value="各商品支出汇总 (未完成订单)").font = bold_font
    current_row += 1
    exp_headers = ["商品编号", "商品名称", "未完成订单数量", "未完成订单金额 (支出)"]
    ws.append(exp_headers)
    for cell in ws[current_row]: cell.font = bold_font; cell.alignment = center_align
    current_row += 1

    grand_total_exp_qty, grand_total_exp_amt = 0, 0.0
    for name in sorted_product_names:
        item = product_data_map[name]
        if item['expenditure_qty'] > 0 or item['expenditure_amount'] != 0:
            ws.append([item['prod_id'], name, item['expenditure_qty'], item['expenditure_amount']])
            grand_total_exp_qty += item['expenditure_qty']
            grand_total_exp_amt += item['expenditure_amount']
            current_row += 1

    ws.append(["总计支出", "", grand_total_exp_qty, grand_total_exp_amt])
    for cell in ws[current_row]: cell.font = bold_font
    current_row += 2

    # --- 净总计 ---
    net_qty = grand_total_income_qty - grand_total_exp_qty
    net_amount = grand_total_income_amt + grand_total_exp_amt
    ws.cell(row=current_row, column=1, value="净总计").font = bold_font
    ws.cell(row=current_row, column=3, value=net_qty).font = bold_font
    ws.cell(row=current_row, column=4, value=net_amount).font = bold_font
    
    # --- 格式化 ---
    for col_letter, width in [('A', 25), ('B', 70), ('C', 20), ('D', 25)]:
        ws.column_dimensions[col_letter].width = width
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.column == 3: cell.number_format = '#,##0'
                if cell.column == 4: cell.number_format = '#,##0.00'

def _format_for_detail_dy(df, is_expenditure=False):
    """根据详情页列定义，格式化DataFrame。"""
    if df.empty: return pd.DataFrame(columns=DETAIL_SHEET_COLUMNS_DY)
    
    detail = pd.DataFrame()
    detail['订单编号'] = df.get(DY_COL_MAIN_ORDER_ID)
    detail['订单状态'] = df.get(DY_COL_ORDER_STATUS)
    detail['售后状态'] = df.get(DY_COL_AFTER_SALES_STATUS)
    detail['取消原因'] = df.get(DY_COL_CANCEL_REASON)
    detail['商品编号'] = df.get(DY_COL_PRODUCT_ID).astype(str)
    detail['商品名称'] = df.get(DY_COL_PRODUCT_NAME_DESC)
    detail['商品单价'] = df.get(DY_COL_UNIT_PRICE)
    detail['商品数量'] = df.get(DY_COL_QUANTITY)
    
    payable_amounts = df.get(CALC_COL_PAYABLE, 0.0)
    detail[CALC_COL_PAYABLE] = -payable_amounts if is_expenditure else payable_amounts
    
    detail['订单提交时间'] = df.get(DY_COL_ORDER_SUBMIT_TIME)
    detail['支付完成时间'] = df.get(DY_COL_PAY_COMPLETE_TIME)
    detail['订单完成时间'] = df.get(DY_COL_ORDER_COMPLETE_TIME)
    
    return detail.reindex(columns=DETAIL_SHEET_COLUMNS_DY).fillna('')

def _create_detail_sheets(wb, product_data_map):
    """为每个商品创建并填充详情页。"""
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')

    for name in sorted(product_data_map.keys()):
        item = product_data_map[name]
        
        # --- Sheet页命名逻辑 ---
        # 1. 直接使用商品名称(name)作为基础，并清理Excel不支持的特殊字符
        base_name = re.sub(r'[\\/\*\[\]\:?]', '_', name)
        
        # 2. 如果清理后的名称长度超过31个字符，则从尾部截取
        if len(base_name) > 31:
            base_name = base_name[-31:] # 直接保留最后31个字符

        # 3. 防重名处理：如果名称已存在，则添加序号
        sheet_name = base_name
        counter = 1
        while sheet_name in wb.sheetnames:
            counter += 1
            suffix = f"({counter})"
            # 为了给序号腾出空间，需要从已截断的base_name上再次截断
            truncated_base = base_name[:31 - len(suffix)]
            sheet_name = f"{truncated_base}{suffix}"

        try:
            ws = wb.create_sheet(sheet_name)
        except Exception as e:
            print(f"警告: 创建Sheet页 '{sheet_name}' 失败: {e}。将使用备用名称。")
            ws = wb.create_sheet(f"{item['prod_id']}_detail")

        # 内部函数，用于写入一个数据区域（如收入明细）
        def write_section(df, title, total_title, is_expenditure=False):
            if df.empty: return
            
            formatted_df = _format_for_detail_dy(df, is_expenditure)
            
            ws.append([title]); ws[ws.max_row][0].font = bold_font
            ws.append(DETAIL_SHEET_COLUMNS_DY)
            for cell in ws[ws.max_row]: cell.font = bold_font; cell.alignment = center_align
            for r in formatted_df.itertuples(index=False): ws.append(list(r))
            
            qty_sum = df[DY_COL_QUANTITY].sum()
            amt_sum = df[CALC_COL_PAYABLE].sum()
            if is_expenditure: amt_sum = -amt_sum
            
            total_row_data = [total_title, "", "", "", "", "", "", qty_sum, amt_sum]
            ws.append(total_row_data)
            for cell in ws[ws.max_row]: cell.font = bold_font
            ws.append([])

        # 写入收入和支出明细
        write_section(item['detail_income_df'], "收入明细 (全部订单)", "收入总计", is_expenditure=False)
        write_section(item['detail_expenditure_df'], "支出明细 (未完成订单)", "支出总计", is_expenditure=True)

        # 设置列宽
        for i, width in enumerate([25, 15, 20, 30, 25, 70, 15, 15, 15, 22, 22, 22], 1):
             ws.column_dimensions[get_column_letter(i)].width = width

# --- 主处理函数 ---

def process_douyin_data(df_raw):
    """
    处理抖店订单DataFrame，生成包含总结和明细的Excel Workbook对象。
    
    Args:
        df_raw (pd.DataFrame): 从CSV文件读取的原始DataFrame。
        
    Returns:
        openpyxl.Workbook: 包含处理结果的Workbook对象，如果处理失败则返回None。
    """
    if df_raw is None or df_raw.empty:
        print("抖店处理错误：输入的DataFrame为空。")
        return None

    # 1. 数据准备与验证
    df_processed = _prepare_and_validate_data(df_raw)
    if df_processed is None: return None

    # 2. 按商品聚合数据
    product_data_map = _aggregate_product_data(df_processed)
    
    # 3. 创建Excel工作簿并生成内容
    wb = Workbook()
    _create_summary_sheet(wb, product_data_map)
    _create_detail_sheets(wb, product_data_map)
    
    # 确保总结页在第一个
    if "销售总结" in wb.sheetnames and wb.sheetnames[0] != "销售总结":
        wb.move_sheet("销售总结", -len(wb.sheetnames) + 1)
        
    return wb

# ---- 主程序入口 (用于独立测试) ----
if __name__ == "__main__":
    # --- 测试配置 ---
    TEST_INPUT_DIR = r"C:\Users\LENOVO\Desktop"
    TEST_OUTPUT_DIR = r"C:\Users\LENOVO\Desktop"
    TEST_FILENAME = "抖店202504月.csv"
    # --- 测试配置结束 ---

    input_file = os.path.join(TEST_INPUT_DIR, TEST_FILENAME)

    if not os.path.exists(input_file):
        print(f"错误: 测试输入文件未找到于 '{os.path.abspath(input_file)}'")
    else:
        # 1. 在测试时，模拟main_processor的行为：先读取文件
        print(f"--- 独立测试: 读取文件 {TEST_FILENAME} ---")
        try:
            # 同样进行基础清洗，模拟main模块可能做的预处理
            df_test_raw = pd.read_csv(input_file, dtype=str, keep_default_na=True, encoding='utf-8-sig')
            df_test_raw.columns = [col.strip().replace('"', '') for col in df_test_raw.columns]
            for col in df_test_raw.columns:
                if df_test_raw[col].dtype == 'object':
                    df_test_raw[col] = df_test_raw[col].astype(str).str.strip().str.replace('\t', '')
                    df_test_raw[col] = df_test_raw[col].replace(
                        ['-', '--', '', 'None', 'nan', '#NULL!', None, 'null'], np.nan, regex=False
                    )
        except Exception as e:
            print(f"测试中读取文件失败: {e}")
            df_test_raw = None

        if df_test_raw is not None:
            # 2. 调用处理函数，传入DataFrame
            print("--- 独立测试: 调用 process_douyin_data ---")
            workbook_result = process_douyin_data(df_test_raw)
            
            # 3. 在测试时，模拟main_processor的行为：保存文件
            if workbook_result:
                os.makedirs(TEST_OUTPUT_DIR, exist_ok=True)
                file_name_without_ext = os.path.splitext(TEST_FILENAME)[0]
                output_filename = f"DY_output_{file_name_without_ext}.xlsx"
                output_file_path = os.path.join(TEST_OUTPUT_DIR, output_filename)
                
                try:
                    workbook_result.save(output_file_path)
                    print(f"\n脚本独立测试成功。输出文件位于: {output_file_path}")
                except Exception as e:
                    print(f"\n测试中保存文件失败: {e}")
            else:
                print("\n脚本独立测试失败，未能生成Workbook对象。")